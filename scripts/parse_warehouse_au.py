# -*- coding: utf-8 -*-
"""解析澳洲海外仓报价 Excel → data/warehouse_au/*.json"""
import json
import os
import re
import sys
import html as _html
import openpyxl

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(
    BASE_DIR,
    "璟川森特（澳洲海外仓）报价-2026.5. 10日（Aupost : iMile 3kg以内实重计费）.xlsx",
)
OUT_DIR = os.path.join(BASE_DIR, "data", "warehouse_au")

SHEET_KEY_MAP = {
    "目录": "mulu",
    "出入库操作费与仓租": "churuku",
    "增值服务": "zengzhi",
    "FBA退货换标": "fba",
    "PFL派送": "pfl",
    "iMile价格表": "imile",
    "AU POST-Eparcel价格表": "aupost_eparcel",
    "YTO价格表": "yto",
    "Fastway价格表": "fastway",
    "Allied价格表": "allied",
    "Border价格表": "border",
    "TFM价格表": "tfm",
    "TOLL价格表": "toll",
    "AU POST-Express价格表": "aupost_express",
    "AU POST- Letter价格表": "aupost_letter",
    "渠道赔偿标准": "peichang",
}

LARGE_THRESHOLD = 500

# Zone abbreviation → price table row name
ZONE_NAME_MAP = {
    "SYD Metro": "Sydney Metro",
    "MEL Metro": "Melbourne Metro",
    "BNE Metro": "Brisbane Metro",
    "ADL Metro": "Adelaide Metro",
    "PER Metro": "Perth Metro",
}


def _cell_val(v):
    if v is None:
        return ""
    if isinstance(v, float):
        if v == int(v):
            return int(v)
        return round(v, 4)
    return str(v).strip()


def _row_empty(row):
    return all(v is None or str(v).strip() == "" for v in row)


def _strip_trailing_empty(rows):
    while rows and all(c == "" for c in rows[-1]):
        rows.pop()
    return rows


def _strip_trailing_empty_cols(headers, rows):
    """Remove trailing empty columns."""
    if not headers:
        return headers, rows
    while headers and headers[-1] == "":
        has_data = any(r[len(headers) - 1] != "" for r in rows if len(r) >= len(headers))
        if not has_data:
            headers.pop()
            for r in rows:
                if len(r) > len(headers):
                    r.pop()
        else:
            break
    return headers, rows


def _is_numeric(v):
    if isinstance(v, (int, float)):
        return True
    try:
        float(str(v))
        return True
    except (ValueError, TypeError):
        return False


def parse_peichang(ws):
    """Parse 渠道赔偿标准 sheet: multiple tables separated by blank rows, with 2-row merged headers."""
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))
    max_col = 7

    sections = []
    title = ""

    # Find title row (row 0 or 1, single text spanning)
    start = 0
    for ri, raw in enumerate(all_rows[:3]):
        vals = [_cell_val(raw[c]) if c < len(raw) else "" for c in range(max_col)]
        non_empty = [v for v in vals if v != ""]
        if len(non_empty) == 1 and len(str(non_empty[0])) > 10:
            title = str(non_empty[0]).replace("璟川森特", "中海豹")
            start = ri + 1
            break

    # Split into blocks by blank rows
    blocks = []
    cur_block = []
    for ri in range(start, len(all_rows)):
        raw = all_rows[ri]
        vals = [_cell_val(raw[c]) if c < len(raw) else "" for c in range(max_col)]
        if all(v == "" for v in vals):
            if cur_block:
                blocks.append(cur_block)
                cur_block = []
        else:
            cur_block.append(vals)
    if cur_block:
        blocks.append(cur_block)

    for bi, block in enumerate(blocks):
        if not block:
            continue

        # Check if first row is a full-span note (like 特别提示)
        first_non_empty = [v for v in block[0] if v != ""]
        if len(block) == 1 and len(first_non_empty) == 1 and len(str(first_non_empty[0])) > 20:
            sections.append({
                "title": "",
                "type": "richtext",
                "html": _html.escape(str(first_non_empty[0])),
            })
            continue

        # Detect header rows: rows where all non-empty values are text (no numbers)
        # and contain header-like keywords
        header_rows = 0
        for row in block[:2]:
            has_num = any(isinstance(v, (int, float)) for v in row if v != "")
            if not has_num:
                header_rows += 1
            else:
                break

        if header_rows == 0:
            header_rows = 1

        # Build merged headers from header rows
        headers = [""] * max_col
        for hi in range(header_rows):
            for ci, v in enumerate(block[hi]):
                if v != "":
                    if headers[ci] == "":
                        headers[ci] = str(v)
                    elif str(v) not in headers[ci]:
                        headers[ci] = str(v)

        # Trim trailing empty headers
        while headers and headers[-1] == "":
            headers.pop()
        col_count = len(headers)

        # Data rows
        data_rows = []
        for row in block[header_rows:]:
            r = row[:col_count]
            while len(r) < col_count:
                r.append("")
            # Check if last row is a full-span note
            non_empty = [v for v in r if v != ""]
            if len(non_empty) == 1 and len(str(non_empty[0])) > 30:
                # This is a note row, add as richtext after this section
                sections.append({
                    "title": "",
                    "headers": headers,
                    "rows": data_rows,
                    "notes": [],
                })
                sections.append({
                    "title": "",
                    "type": "richtext",
                    "html": _html.escape(str(non_empty[0])),
                })
                data_rows = None
                break
            data_rows.append(r)

        if data_rows is not None and data_rows:
            sec_title = ""
            if bi == 0 and title:
                sec_title = title
            sections.append({
                "title": sec_title,
                "headers": headers,
                "rows": data_rows,
                "notes": [],
            })

    return sections


def parse_aupost_letter(ws):
    """Parse AU POST Letter sheet: title row, header row, data rows, notes."""
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))
    sections = []
    headers = None
    data_rows = []
    notes = []
    title = ""

    for ri, raw in enumerate(all_rows):
        vals = [_cell_val(raw[c]) if c < len(raw) else "" for c in range(min(len(raw), 10))]
        # Trim trailing empties
        while vals and vals[-1] == "":
            vals.pop()
        if not vals:
            continue

        non_empty = [(i, v) for i, v in enumerate(vals) if v != ""]
        if not non_empty:
            continue

        # Title row (long text spanning the row)
        if not title and len(non_empty) == 1 and len(str(non_empty[0][1])) > 20:
            title = str(non_empty[0][1])
            continue

        # Header row (multiple non-numeric text values)
        if headers is None and len(non_empty) >= 3 and all(not isinstance(v, (int, float)) for _, v in non_empty):
            headers = vals
            continue

        # Notes rows (single text value, appears after data)
        if headers and not any(isinstance(v, (int, float)) for _, v in non_empty) and data_rows:
            notes.append(" ".join(str(v) for _, v in non_empty))
            continue

        # Data row
        if headers:
            row = vals
            while len(row) < len(headers):
                row.append("")
            data_rows.append(row[:len(headers)])

    if headers and data_rows:
        sections.append({
            "title": title,
            "type": "price_table",
            "headers": headers,
            "rows": data_rows,
            "notes": [],
        })

    if notes:
        sections.append({
            "title": "价格说明",
            "type": "richtext",
            "html": "<br>".join(_html.escape(n) for n in notes),
        })

    return sections


def parse_simple_table(ws, skip_rows=0, max_col=None):
    """Parse a simple sheet into sections split by blank rows."""
    all_rows = []
    for row in ws.iter_rows(min_row=1 + skip_rows, max_col=max_col, values_only=True):
        all_rows.append([_cell_val(c) for c in row])

    _strip_trailing_empty(all_rows)

    sections = []
    current_section = {"title": "", "rows": []}

    for row in all_rows:
        if all(c == "" for c in row):
            if current_section["rows"]:
                sections.append(current_section)
                current_section = {"title": "", "rows": []}
            continue
        current_section["rows"].append(row)

    if current_section["rows"]:
        sections.append(current_section)

    for sec in sections:
        if sec["rows"]:
            first = sec["rows"][0]
            non_empty = [c for c in first if c != ""]
            if len(non_empty) == 1 and not _is_numeric(non_empty[0]):
                sec["title"] = str(non_empty[0])
                sec["rows"] = sec["rows"][1:]

        rows = sec["rows"]
        if rows:
            max_c = max(len(r) for r in rows)
            for r in rows:
                while len(r) < max_c:
                    r.append("")
            while max_c > 0 and all(r[max_c - 1] == "" for r in rows):
                max_c -= 1
                for r in rows:
                    r.pop()

    sections = [s for s in sections if s["rows"]]

    for sec in sections:
        if sec.get("title") == "报价说明":
            lines = []
            for row in sec["rows"]:
                text = " ".join(str(c) for c in row if c != "")
                if text:
                    lines.append(text)
            sec["type"] = "richtext"
            sec["html"] = "".join("<p>" + l + "</p>" for l in lines)
            del sec["rows"]

    return sections


def parse_mulu(ws):
    """Parse the 目录 (table of contents) sheet."""
    rows = []
    for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
        v = _cell_val(row[0])
        if v:
            rows.append(v)
    info_lines = []
    nav_items = []
    started_nav = False
    for v in rows:
        cleaned = v.replace("\n", "").strip()
        if "服务类目" in cleaned:
            started_nav = True
            continue
        if not started_nav:
            info_lines.append(cleaned)
        else:
            nav_items.append(cleaned)

    sections = []
    if info_lines:
        sections.append({"title": "公司信息", "type": "info", "rows": [[l] for l in info_lines]})
    if nav_items:
        sections.append({"title": "服务类目", "type": "nav", "rows": [[n] for n in nav_items]})
    return sections


def parse_pricing_with_zones(ws, sheet_name):
    """Parse large sheets that have price table on the left and zone table on the right."""
    all_rows_raw = list(ws.iter_rows(min_row=1, values_only=True))
    max_col = ws.max_column

    zone_start_col = None
    for row in all_rows_raw[:10]:
        for ci, v in enumerate(row):
            s = str(v or "").strip()
            if s in ("Zone", "PFL分区", "分区") and ci > 5:
                zone_start_col = ci
                break
            if "分区" in s and ci > 5:
                zone_start_col = ci
                break
        if zone_start_col is not None:
            break

    if zone_start_col and zone_start_col > 0:
        gap_col = zone_start_col - 1
        left_end = gap_col
    else:
        left_end = max_col
        zone_start_col = None

    price_sections = _parse_price_table_half(all_rows_raw, 0, left_end)

    zone_section = None
    if zone_start_col is not None:
        zone_section = _parse_zone_table_half(all_rows_raw, zone_start_col, max_col)

    return price_sections, zone_section


def _parse_price_table_half(all_rows_raw, col_start, col_end):
    """Parse the left-side price table portion of a sheet."""
    rows = []
    for raw in all_rows_raw:
        row = [_cell_val(raw[c]) if c < len(raw) else "" for c in range(col_start, col_end)]
        rows.append(row)

    _strip_trailing_empty(rows)

    sections = []
    current = {"title": "", "headers": [], "rows": [], "notes": []}
    in_notes = False

    for row in rows:
        if all(c == "" for c in row):
            if current["headers"] or current["rows"]:
                sections.append(current)
                current = {"title": "", "headers": [], "rows": [], "notes": []}
                in_notes = False
            continue

        non_empty = [(i, c) for i, c in enumerate(row) if c != ""]

        if not current["headers"] and not current["rows"]:
            if len(non_empty) <= 2 and all(not _is_numeric(c) for _, c in non_empty):
                title_parts = [str(c) for _, c in non_empty]
                new_title = " ".join(title_parts)
                if current["title"]:
                    # Extract parenthesized prefix like （悉尼仓） from previous title
                    import re as _re
                    paren = _re.match(r"(（[^）]+）)", current["title"])
                    if paren:
                        current["title"] = paren.group(1) + new_title
                    else:
                        current["title"] += " — " + new_title
                else:
                    current["title"] = new_title
                continue

        if not current["headers"] and len(non_empty) >= 3:
            current["headers"] = row
            continue

        if current["headers"]:
            has_numbers = any(_is_numeric(c) for c in row if c != "")
            if has_numbers and not in_notes:
                current["rows"].append(row)
            else:
                in_notes = True
                text = " ".join(str(c) for c in row if c != "")
                if text:
                    current["notes"].append(text)
        else:
            text = " ".join(str(c) for c in row if c != "")
            if text:
                current["notes"].append(text)

    if current["headers"] or current["rows"] or current["notes"]:
        sections.append(current)

    for sec in sections:
        if sec["rows"] and sec["headers"]:
            h_len = len(sec["headers"])
            for r in sec["rows"]:
                while len(r) < h_len:
                    r.append("")
                while len(r) > h_len:
                    r.pop()

    return [s for s in sections if s.get("headers") or s.get("rows") or s.get("notes")]


def _parse_zone_table_half(all_rows_raw, col_start, col_end):
    """Parse the right-side zone/area table."""
    headers = []
    rows = []
    for raw in all_rows_raw:
        row = [_cell_val(raw[c]) if c < len(raw) else "" for c in range(col_start, col_end)]
        if all(c == "" for c in row):
            continue
        if not headers:
            if any(str(c).strip() for c in row):
                headers = row
                continue
        rows.append(row)

    # Determine actual column count from data rows (some sheets have
    # a short title-row as the first header, e.g. ["PFL分区","","",""],
    # while the real multi-column data starts later).
    max_data_cols = len(headers)
    for r in rows:
        non_trail = len(r)
        while non_trail > 0 and r[non_trail - 1] == "":
            non_trail -= 1
        if non_trail > max_data_cols:
            max_data_cols = non_trail

    # Trim headers to max_data_cols (strip trailing empties beyond that)
    while len(headers) > max_data_cols:
        headers.pop()
    while len(headers) < max_data_cols:
        headers.append("")

    for r in rows:
        while len(r) > max_data_cols:
            r.pop()
        while len(r) < max_data_cols:
            r.append("")

    rows = [r for r in rows if any(c != "" for c in r)]

    if not headers and not rows:
        return None

    return {
        "title": "分区表",
        "type": "zone_table",
        "headers": headers,
        "rows": rows,
    }


def parse_imile(ws, sheet_name):
    """Parse iMile sheet which has a unique layout:

    - Two price tables stacked vertically (Sydney warehouse, Melbourne warehouse)
    - Each has a 2-row merged header (row1: Origin/Zone/实重计费/3<W≤5KG...;
      row2: 0<W≤0.5KG/0.5<W≤1KG/1<W≤2KG/2<W≤3KG)
    - Notes in col 11 alongside price rows
    - Two zone tables side-by-side BELOW the price tables (Sydney cols 0-3,
      Melbourne cols 6-9)
    """
    all_rows_raw = list(ws.iter_rows(min_row=1, values_only=True))

    # --- Step 1: find the boundary between price area and zone area ---
    zone_boundary = None
    for ri, raw in enumerate(all_rows_raw):
        s = str(raw[0] or "").strip()
        if "分区" in s and "发出" in s:
            zone_boundary = ri
            break

    if zone_boundary is None:
        zone_boundary = len(all_rows_raw)

    # --- Step 2: parse price tables from the upper portion ---
    price_sections = []
    current = None  # {title, headers, rows, notes}

    for ri in range(zone_boundary):
        raw = all_rows_raw[ri]
        # Convert to clean values (cols 0-9 for price, col 11 for notes)
        price_cols = [_cell_val(raw[c]) if c < len(raw) else "" for c in range(10)]
        note_val = _cell_val(raw[11]) if len(raw) > 11 else ""

        all_empty = all(c == "" for c in price_cols)
        if all_empty:
            # Blank row — finalize current section
            if current and (current["rows"] or current["headers"]):
                price_sections.append(current)
                current = None
            continue

        non_empty = [(i, c) for i, c in enumerate(price_cols) if c != ""]

        # Detect title row: 1-2 non-empty text cells, no numbers
        if len(non_empty) <= 2 and all(not _is_numeric(c) for _, c in non_empty):
            has_data_below = False
            for rj in range(ri + 1, min(ri + 4, zone_boundary)):
                rr = all_rows_raw[rj]
                if any(_is_numeric(rr[c]) for c in range(10) if c < len(rr) and rr[c] is not None):
                    has_data_below = True
                    break
            if has_data_below or current is None:
                # Start new section
                if current and (current["rows"] or current["headers"]):
                    price_sections.append(current)
                title = " ".join(str(c) for _, c in non_empty)
                current = {"title": title, "headers": [], "rows": [], "notes": []}
                continue

        if current is None:
            current = {"title": "", "headers": [], "rows": [], "notes": []}

        # Detect header row: has "Origin" or "Zone" or "实重计费" text
        texts = [str(c).strip() for _, c in non_empty]
        is_header_row = any(t in ("Origin", "Zone", "实重计费") for t in texts)

        # Detect sub-header row: fills in the gaps of the main header (cols 2-5)
        # In iMile, row 3 has "实重计费" spanning cols 2-5, row 4 has actual weight
        # range names like "0＜W≤0.5KG" that should overwrite.
        is_subheader = False
        if current["headers"] and not current["rows"]:
            cur_h = current["headers"]
            fills_gap = False
            for i, c in non_empty:
                if i < len(cur_h) and not _is_numeric(c):
                    # Either fills an empty slot, or overwrites a group label
                    if cur_h[i] == "":
                        fills_gap = True
                        break
                    # Check if this is a weight range overwriting a group label
                    cs = str(c)
                    if ("W" in cs or "KG" in cs or "kg" in cs) and cs != cur_h[i]:
                        fills_gap = True
                        break
            if fills_gap:
                is_subheader = True

        if is_header_row and not current["headers"]:
            current["headers"] = list(price_cols)
            continue

        if is_subheader:
            # Merge sub-header into existing header — overwrite group labels too
            for i, c in non_empty:
                if i < len(current["headers"]) and not _is_numeric(c):
                    current["headers"][i] = c
            # Also pick up notes from sub-header row
            if note_val:
                current["notes"].append(str(note_val).replace("\n", " ").strip())
            continue

        # Data row — has numeric values
        has_numbers = any(_is_numeric(c) for _, c in non_empty)
        if has_numbers:
            current["rows"].append(list(price_cols))
            if note_val:
                current["notes"].append(str(note_val).replace("\n", " ").strip())
            continue

        # Fallback: text-only row, treat as note
        text = " ".join(str(c) for _, c in non_empty)
        if text:
            current["notes"].append(text)
        if note_val:
            current["notes"].append(str(note_val).replace("\n", " ").strip())

    if current and (current["rows"] or current["headers"]):
        price_sections.append(current)

    # Clean up headers — strip trailing empty columns
    for sec in price_sections:
        if sec["headers"]:
            h = sec["headers"]
            while h and h[-1] == "":
                h.pop()
            for r in sec["rows"]:
                while len(r) > len(h):
                    r.pop()
                while len(r) < len(h):
                    r.append("")

    # --- Step 3: parse zone tables from the lower portion ---
    # Two side-by-side zone tables: Sydney (cols 0-3), Melbourne (cols 6-9)
    zone_rows_raw = all_rows_raw[zone_boundary:]

    # Find the header row with "Suburb", "Postcode", etc.
    zone_header_ri = None
    for ri, raw in enumerate(zone_rows_raw):
        for ci in range(4):
            s = str(raw[ci] or "").strip().lower() if ci < len(raw) else ""
            if s in ("suburb", "postcode"):
                zone_header_ri = ri
                break
        if zone_header_ri is not None:
            break

    zone_section = None
    postcode_zone_map = None
    if zone_header_ri is not None:
        # Build zone section (Sydney side, cols 0-3) for display
        z_headers = [_cell_val(zone_rows_raw[zone_header_ri][c])
                     if c < len(zone_rows_raw[zone_header_ri]) else ""
                     for c in range(4)]
        z_rows = []
        for raw in zone_rows_raw[zone_header_ri + 1:]:
            row = [_cell_val(raw[c]) if c < len(raw) else "" for c in range(4)]
            if any(c != "" for c in row):
                z_rows.append(row)

        if z_rows:
            zone_section = {
                "title": "分区表",
                "type": "zone_table",
                "headers": z_headers,
                "rows": z_rows,
            }

        # Build separate postcode → zone maps for Sydney and Melbourne
        postcode_zone_maps = {}
        map_keys = {0: "sydney", 6: "melbourne"}

        for col_offset, map_key in map_keys.items():
            pc_col = None
            zn_col = None
            hdr_raw = zone_rows_raw[zone_header_ri]
            for ci in range(col_offset, col_offset + 4):
                s = str(hdr_raw[ci] or "").strip().lower() if ci < len(hdr_raw) else ""
                if s == "postcode":
                    pc_col = ci
                elif s == "zone":
                    zn_col = ci
            if pc_col is None or zn_col is None:
                continue

            mapping = {}
            for raw in zone_rows_raw[zone_header_ri + 1:]:
                pc_raw = str(raw[pc_col] or "").strip() if pc_col < len(raw) else ""
                zn_raw = str(raw[zn_col] or "").strip() if zn_col < len(raw) else ""
                if not pc_raw or not zn_raw:
                    continue
                digits = "".join(c for c in pc_raw if c.isdigit())
                if len(digits) != 4:
                    continue
                zone_name = ZONE_NAME_MAP.get(zn_raw, zn_raw)
                mapping[digits] = zone_name
            if mapping:
                postcode_zone_maps[map_key] = mapping

        if not postcode_zone_maps:
            postcode_zone_maps = None

    return price_sections, zone_section, postcode_zone_maps


def parse_aupost_dual(ws, sheet_name):
    """Parse AU POST sheets with side-by-side Sydney/Melbourne price tables.

    Layout: Sydney (cols 0-15) | gap (col 16) | Melbourne (cols 17-32)
    Below: notes, then a shared zone table (cols 0-3).
    Headers span 3 rows (merged cells for weight ranges + Base/Kilo).
    """
    all_rows_raw = list(ws.iter_rows(min_row=1, values_only=True))
    max_col = ws.max_column

    # --- Step 1: find title row (contains warehouse labels) ---
    title_row_idx = None
    syd_title = ""
    mel_title = ""
    for ri, raw in enumerate(all_rows_raw):
        s0 = str(raw[0] or "").strip()
        if "悉尼" in s0 or "墨尔本" in s0:
            title_row_idx = ri
            # Collect title parts from both sides
            syd_parts = []
            mel_parts = []
            for ci, v in enumerate(raw):
                if v is None:
                    continue
                vs = str(v).strip()
                if not vs:
                    continue
                if ci < 16:
                    syd_parts.append(vs)
                elif ci >= 17:
                    mel_parts.append(vs)
            syd_title = " ".join(syd_parts)
            mel_title = " ".join(mel_parts)
            break

    if title_row_idx is None:
        title_row_idx = 0

    # --- Step 2: find the 3-row merged header ---
    # Look for rows after title that have header-like text (State, Zone Code, etc.)
    header_start = title_row_idx + 1
    header_rows_raw = all_rows_raw[header_start:header_start + 4]

    # Build merged headers for Sydney (cols 0-15) and Melbourne (cols 17-32)
    syd_col_count = 16
    mel_col_count = 16
    mel_offset = 17

    def _build_merged_header(col_start, col_count, header_rows):
        header = [""] * col_count
        for row_raw in header_rows:
            has_numeric = False
            for ci in range(col_count):
                abs_ci = col_start + ci
                v = row_raw[abs_ci] if abs_ci < len(row_raw) else None
                if isinstance(v, (int, float)):
                    has_numeric = True
                    break
            if has_numeric:
                break
            for ci in range(col_count):
                abs_ci = col_start + ci
                v = row_raw[abs_ci] if abs_ci < len(row_raw) and row_raw[abs_ci] else ""
                vs = str(v).strip() if v else ""
                if not vs:
                    continue
                if not header[ci]:
                    header[ci] = vs
                else:
                    header[ci] = vs
        return header

    syd_header = _build_merged_header(0, syd_col_count, header_rows_raw)
    mel_header = _build_merged_header(mel_offset, mel_col_count, header_rows_raw)

    # --- Step 3: find data rows (first row with numeric data after headers) ---
    data_start = None
    for ri in range(header_start, min(header_start + 6, len(all_rows_raw))):
        raw = all_rows_raw[ri]
        has_num = any(isinstance(raw[c], (int, float)) for c in range(syd_col_count) if c < len(raw) and raw[c] is not None)
        if has_num:
            data_start = ri
            break

    if data_start is None:
        data_start = header_start + 3

    # Find end of price data: first fully-empty row or text-only row after data
    data_end = data_start
    for ri in range(data_start, len(all_rows_raw)):
        raw = all_rows_raw[ri]
        syd_vals = [raw[c] for c in range(syd_col_count) if c < len(raw)]
        has_num = any(isinstance(v, (int, float)) for v in syd_vals if v is not None)
        if not has_num:
            data_end = ri
            break
    else:
        data_end = len(all_rows_raw)

    # Extract price rows
    def _extract_rows(col_start, col_count, row_start, row_end):
        rows = []
        for ri in range(row_start, row_end):
            raw = all_rows_raw[ri]
            row = [_cell_val(raw[col_start + c]) if (col_start + c) < len(raw) else ""
                   for c in range(col_count)]
            if any(c != "" for c in row):
                rows.append(row)
        return rows

    syd_rows = _extract_rows(0, syd_col_count, data_start, data_end)
    mel_rows = _extract_rows(mel_offset, mel_col_count, data_start, data_end)

    # --- Step 4: extract notes (text rows between price data and zone table) ---
    notes = []
    zone_boundary = len(all_rows_raw)
    for ri in range(data_end, len(all_rows_raw)):
        raw = all_rows_raw[ri]
        s0 = str(raw[0] or "").strip()
        if s0 == "分区表" or s0 == "分区":
            zone_boundary = ri
            break
        # Zone table header row (Suburb, Postcode, etc.)
        if s0.lower() in ("suburb", "postcode", "postcodes"):
            zone_boundary = ri
            break
        if s0:
            notes.append(s0)

    # --- Step 5: build price sections ---
    price_sections = []

    syd_sec = {
        "title": syd_title,
        "headers": syd_header,
        "rows": syd_rows,
        "notes": [],
    }
    price_sections.append(syd_sec)

    mel_sec = {
        "title": mel_title,
        "headers": mel_header,
        "rows": mel_rows,
        "notes": [],
    }
    price_sections.append(mel_sec)

    if notes:
        price_sections.append({
            "title": "价格说明",
            "type": "richtext",
            "html": "<br>".join(_html.escape(n) for n in notes),
        })

    # Strip trailing empty columns from headers/rows
    for sec in price_sections:
        if sec.get("type") == "richtext":
            continue
        h = sec["headers"]
        while h and h[-1] == "":
            h.pop()
        for r in sec["rows"]:
            while len(r) > len(h):
                r.pop()
            while len(r) < len(h):
                r.append("")

    # --- Step 6: parse zone table ---
    zone_section = None
    postcode_zone_map = None

    zone_rows_raw = all_rows_raw[zone_boundary:]
    zone_header_ri = None
    for ri, raw in enumerate(zone_rows_raw):
        for ci in range(4):
            s = str(raw[ci] or "").strip().lower() if ci < len(raw) else ""
            if s in ("suburb", "postcode", "postcodes"):
                zone_header_ri = ri
                break
        if zone_header_ri is not None:
            break

    if zone_header_ri is not None:
        z_headers = [_cell_val(zone_rows_raw[zone_header_ri][c])
                     if c < len(zone_rows_raw[zone_header_ri]) else ""
                     for c in range(4)]
        z_rows = []
        for raw in zone_rows_raw[zone_header_ri + 1:]:
            row = [_cell_val(raw[c]) if c < len(raw) else "" for c in range(4)]
            if any(c != "" for c in row):
                z_rows.append(row)

        if z_rows:
            zone_section = {
                "title": "分区表",
                "type": "zone_table",
                "headers": z_headers,
                "rows": z_rows,
            }

        # Build postcode → zone map
        pc_col = None
        zn_col = None
        for ci, h in enumerate(z_headers):
            hl = str(h).strip().lower()
            if hl in ("postcode", "postcodes", "邮编"):
                pc_col = ci
            elif hl in ("zone", "分区"):
                zn_col = ci

        if pc_col is not None and zn_col is not None:
            postcode_zone_map = {}
            for row in z_rows:
                if len(row) <= max(pc_col, zn_col):
                    continue
                pc_raw = str(row[pc_col]).strip()
                zn_raw = str(row[zn_col]).strip()
                if not pc_raw or not zn_raw:
                    continue
                digits = "".join(c for c in pc_raw if c.isdigit())
                if len(digits) < 3 or len(digits) > 4:
                    continue
                digits = digits.zfill(4)
                zone_name = ZONE_NAME_MAP.get(zn_raw, zn_raw)
                postcode_zone_map[digits] = zone_name
            if not postcode_zone_map:
                postcode_zone_map = None

    return price_sections, zone_section, postcode_zone_map


def parse_generic_large(ws, sheet_name):
    """Parse large sheets that don't have a separate zone table on the right."""
    all_rows_raw = list(ws.iter_rows(min_row=1, values_only=True))
    price_sections = _parse_price_table_half(all_rows_raw, 0, ws.max_column)
    return price_sections, None


def parse_yto(ws, sheet_name):
    """Parse YTO: stacked Sydney/Melbourne price tables + notes col 13 + zone table below."""
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))

    # Find zone table start (row with "YTO分区表" or "分区表" in col 0)
    zone_start_row = None
    for ri, raw in enumerate(all_rows):
        v = _cell_val(raw[0]) if raw else ""
        if "分区表" in str(v):
            zone_start_row = ri
            break

    price_end = zone_start_row if zone_start_row else len(all_rows)

    # Parse price tables manually (2-row headers, stacked Sydney/Melbourne)
    sections = []
    notes = []
    ri = 0
    while ri < price_end:
        raw = all_rows[ri]
        # Collect notes from col 12
        v12 = _cell_val(raw[12]) if len(raw) > 12 else ""
        if v12:
            notes.append(str(v12))

        row = [_cell_val(raw[c]) if c < len(raw) else "" for c in range(0, 12)]
        non_empty = [(i, c) for i, c in enumerate(row) if c != ""]

        if not non_empty:
            ri += 1
            continue

        # Detect title row (1-2 non-empty, all non-numeric)
        if len(non_empty) <= 2 and all(not _is_numeric(c) for _, c in non_empty):
            title = " ".join(str(c) for _, c in non_empty)
            ri += 1
            # Next row should be header row 1 (Origin, Zone, 实重计费...)
            if ri >= price_end:
                break
            raw = all_rows[ri]
            if len(raw) > 12:
                v12 = _cell_val(raw[12])
                if v12:
                    notes.append(str(v12))
            hdr1 = [_cell_val(raw[c]) if c < len(raw) else "" for c in range(0, 12)]
            ri += 1
            # Next row = header row 2 (weight ranges)
            if ri >= price_end:
                break
            raw = all_rows[ri]
            if len(raw) > 12:
                v12 = _cell_val(raw[12])
                if v12:
                    notes.append(str(v12))
            hdr2 = [_cell_val(raw[c]) if c < len(raw) else "" for c in range(0, 12)]
            ri += 1

            # Merge headers: hdr2 fills in blanks of hdr1
            headers = []
            for ci in range(12):
                h1 = hdr1[ci] if ci < len(hdr1) else ""
                h2 = hdr2[ci] if ci < len(hdr2) else ""
                headers.append(str(h2) if h2 else str(h1))

            # Read data rows until blank row
            data_rows = []
            while ri < price_end:
                raw = all_rows[ri]
                if len(raw) > 12:
                    v12 = _cell_val(raw[12])
                    if v12:
                        notes.append(str(v12))
                row = [_cell_val(raw[c]) if c < len(raw) else "" for c in range(0, 12)]
                if all(c == "" for c in row):
                    ri += 1
                    break
                if any(_is_numeric(c) for c in row if c != ""):
                    data_rows.append(row)
                ri += 1

            # Strip trailing empty cols
            headers, data_rows = _strip_trailing_empty_cols(headers, data_rows)
            sections.append({"title": title, "headers": headers, "rows": data_rows})
        else:
            ri += 1

    if notes:
        sections.append({"title": "价格说明", "type": "richtext", "html": "<br>".join(_html.escape(n) for n in notes)})

    # Parse zone table
    zone_section = None
    if zone_start_row:
        zone_rows = all_rows[zone_start_row:]
        zone_section = _parse_zone_from_rows(zone_rows, 0, 4)

    return sections, zone_section


def parse_fastway(ws, sheet_name):
    """Parse Fastway: stacked 1-row price tables (2-row headers) + notes + zone on right (cols 13-16)."""
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))

    # Price tables on left (cols 0-11), same 2-row header pattern as YTO
    sections = []
    notes = []
    ri = 0
    max_price_col = 12

    while ri < len(all_rows):
        raw = all_rows[ri]
        row = [_cell_val(raw[c]) if c < len(raw) else "" for c in range(0, max_price_col)]
        non_empty = [(i, c) for i, c in enumerate(row) if c != ""]

        if not non_empty:
            ri += 1
            continue

        # Title row
        if len(non_empty) <= 2 and all(not _is_numeric(c) for _, c in non_empty):
            first_val = str(non_empty[0][1])
            # Check if it's a notes section
            if "价格说明" in first_val or (sections and len(first_val) > 20):
                # Rest is notes
                notes.append(first_val)
                ri += 1
                while ri < len(all_rows):
                    raw = all_rows[ri]
                    row = [_cell_val(raw[c]) if c < len(raw) else "" for c in range(0, max_price_col)]
                    text = " ".join(str(c) for c in row if c != "")
                    if text:
                        notes.append(text)
                    ri += 1
                break

            title = " ".join(str(c) for _, c in non_empty)
            ri += 1
            # Header row 1
            if ri >= len(all_rows):
                break
            raw = all_rows[ri]
            hdr1 = [_cell_val(raw[c]) if c < len(raw) else "" for c in range(0, max_price_col)]
            ri += 1
            # Header row 2
            if ri >= len(all_rows):
                break
            raw = all_rows[ri]
            hdr2 = [_cell_val(raw[c]) if c < len(raw) else "" for c in range(0, max_price_col)]
            ri += 1

            # Merge headers
            headers = []
            for ci in range(max_price_col):
                h1 = hdr1[ci] if ci < len(hdr1) else ""
                h2 = hdr2[ci] if ci < len(hdr2) else ""
                headers.append(str(h2) if h2 else str(h1))

            # Data rows
            data_rows = []
            while ri < len(all_rows):
                raw = all_rows[ri]
                row = [_cell_val(raw[c]) if c < len(raw) else "" for c in range(0, max_price_col)]
                if all(c == "" for c in row):
                    ri += 1
                    break
                if any(_is_numeric(c) for c in row if c != ""):
                    data_rows.append(row)
                    ri += 1
                else:
                    break

            headers, data_rows = _strip_trailing_empty_cols(headers, data_rows)
            sections.append({"title": title, "headers": headers, "rows": data_rows})
        else:
            # Could be notes continuation
            text = " ".join(str(c) for _, c in non_empty)
            if sections and text:
                notes.append(text)
            ri += 1

    if notes:
        sections.append({"title": "价格说明", "type": "richtext", "html": "<br>".join(_html.escape(n) for n in notes)})

    # Zone table on right (cols 13-16: Zone, Suburb, Postcode, State)
    zone_section = _parse_zone_from_rows(all_rows, 13, 17)

    return sections, zone_section


def parse_dual_simple(ws, sheet_name, left_cols, right_cols, notes_cols=None, zone_below=False):
    """Parse side-by-side dual warehouse sheets (Allied, Border, TFM, TOLL).

    Args:
        left_cols: (start, end) for Sydney price columns
        right_cols: (start, end) for Melbourne price columns
        notes_cols: (start, end) for notes/surcharges columns, or None
        zone_below: if True, zone table starts after price data in cols 0-7
    """
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))

    # Find where the AUPOST/zone table marker is
    zone_start_row = None
    for ri, raw in enumerate(all_rows):
        v = _cell_val(raw[0]) if raw else ""
        if "分区表" in str(v) or "AUPOST分区" in str(v):
            zone_start_row = ri
            break

    price_end = zone_start_row if zone_start_row else len(all_rows)
    price_rows = all_rows[:price_end]

    # Parse left (Sydney) price table
    left_sections = _parse_dual_half(price_rows, left_cols[0], left_cols[1], "（悉尼仓）")
    # Parse right (Melbourne) price table
    right_sections = _parse_dual_half(price_rows, right_cols[0], right_cols[1], "（墨尔本仓）")

    sections = left_sections + right_sections

    # Parse notes
    if notes_cols:
        text_lines = []
        table_headers = None
        table_rows = []
        in_table = False
        col_count = notes_cols[1] - notes_cols[0]
        for raw in price_rows:
            cols_data = []
            for c in range(notes_cols[0], min(notes_cols[1], len(raw))):
                v = _cell_val(raw[c]) if c < len(raw) else ""
                cols_data.append(str(v) if v else "")
            non_empty = [v for v in cols_data if v]
            if not non_empty:
                continue
            if not in_table:
                # Table header: first 3 cols all filled (consecutive)
                if len(cols_data) >= 3 and cols_data[0] and cols_data[1] and cols_data[2]:
                    table_headers = cols_data[:3]
                    in_table = True
                else:
                    text_lines.append("  ".join(non_empty) if len(non_empty) > 1 else non_empty[0])
            else:
                if cols_data[0] and cols_data[1]:
                    table_rows.append(cols_data[:len(table_headers)])
                elif len(non_empty) == 1:
                    text_lines.append(non_empty[0])

        if text_lines:
            note_html = "<br>".join(_html.escape(l) for l in text_lines)
            sections.append({"title": "价格说明", "type": "richtext", "html": note_html})
        if table_headers and table_rows:
            sections.append({"title": "附加费", "type": "price_table", "headers": table_headers, "rows": table_rows, "notes": []})

    # Parse zone table below
    zone_section = None
    if zone_below and zone_start_row:
        zone_rows = all_rows[zone_start_row:]
        # Detect col count from header row
        max_zone_col = 7
        for raw in zone_rows[1:5]:
            for c in range(max_zone_col, len(raw)):
                if raw[c] is not None:
                    max_zone_col = c + 1
        zone_section = _parse_zone_from_rows(zone_rows, 0, min(max_zone_col, 8))

    return sections, zone_section


def parse_toll(ws, sheet_name):
    """Parse TOLL: side-by-side price + zone tables on RIGHT (cols 11-14, 15-18)."""
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))

    # Find where price ends (notes start with "价格说明")
    price_end = len(all_rows)
    for ri, raw in enumerate(all_rows):
        v = _cell_val(raw[0]) if raw else ""
        if "价格说明" in str(v):
            price_end = ri
            break

    price_rows = all_rows[:price_end]

    # Left (Sydney): cols 0-3, Right (Melbourne): cols 6-9
    left_sections = _parse_dual_half(price_rows, 0, 4, "（悉尼仓）")
    right_sections = _parse_dual_half(price_rows, 6, 10, "（墨尔本仓）")

    sections = left_sections + right_sections

    # Notes below price
    notes = []
    for ri in range(price_end, len(all_rows)):
        raw = all_rows[ri]
        v = _cell_val(raw[0]) if raw else ""
        if v:
            notes.append(str(v))
    if notes:
        sections.append({"title": "价格说明", "type": "richtext", "html": "<br>".join(_html.escape(n) for n in notes)})

    # Zone table on right: cols 11-14 (Postcode, Suburb, Zone)
    zone_section = _parse_zone_from_rows(all_rows, 11, 15)

    return sections, zone_section


def _parse_dual_half(all_rows, col_start, col_end, warehouse_prefix):
    """Parse one half of a side-by-side dual table into sections."""
    rows = []
    for raw in all_rows:
        row = [_cell_val(raw[c]) if c < len(raw) else "" for c in range(col_start, col_end)]
        rows.append(row)

    _strip_trailing_empty(rows)

    title = ""
    headers = []
    data_rows = []
    state = "title"

    for row in rows:
        if all(c == "" for c in row):
            continue

        non_empty = [(i, c) for i, c in enumerate(row) if c != ""]

        if state == "title":
            if len(non_empty) <= 2 and all(not _is_numeric(c) for _, c in non_empty):
                parts = [str(c) for _, c in non_empty]
                new_part = " ".join(parts)
                if not title:
                    if "悉尼" not in new_part and "墨尔本" not in new_part:
                        title = warehouse_prefix + new_part
                    else:
                        title = new_part
                else:
                    title = title + " " + new_part
                continue
            headers = row
            state = "data"
            continue

        if state == "data":
            if any(_is_numeric(c) for c in row if c != ""):
                data_rows.append(row)
            elif len(non_empty) == 1 and non_empty[0][0] == 0:
                text = str(non_empty[0][1])
                # Notes detection: starts with digit+period or "价格说明"
                if re.match(r"^\d+[.、.]", text) or "价格说明" in text:
                    # This and everything after is notes — stop data parsing
                    break
                # State grouping row (e.g. "New South Wales") — keep as marker
                marker = [""] * len(headers) if headers else [""] * (col_end - col_start)
                marker[0] = text
                data_rows.append(marker)

    if not headers and not data_rows:
        return []

    h_len = len(headers) if headers else (max(len(r) for r in data_rows) if data_rows else 0)
    for r in data_rows:
        while len(r) < h_len:
            r.append("")
        while len(r) > h_len:
            r.pop()

    return [{"title": title, "headers": headers, "rows": data_rows}]


def _parse_zone_from_rows(all_rows, col_start, col_end):
    """Parse a zone table from rows, finding headers and data."""
    headers = []
    data_rows = []
    found_header = False

    for raw in all_rows:
        row = [_cell_val(raw[c]) if c < len(raw) else "" for c in range(col_start, col_end)]
        if all(c == "" for c in row):
            continue

        # Skip title rows (like "YTO分区表", "AUPOST分区表")
        non_empty = [c for c in row if c != ""]
        if len(non_empty) == 1 and not _is_numeric(non_empty[0]) and "分区" in str(non_empty[0]):
            continue

        if not found_header:
            if len(non_empty) >= 2 and all(not _is_numeric(c) for c in non_empty):
                headers = row
                found_header = True
                continue
            if any(_is_numeric(c) for c in non_empty):
                data_rows.append(row)
                found_header = True
                continue
            continue

        data_rows.append(row)

    # Strip trailing empty cols
    if headers:
        headers, data_rows = _strip_trailing_empty_cols(headers, data_rows)

    # Remove columns that are entirely empty (gap columns)
    if headers and data_rows:
        keep = [ci for ci, h in enumerate(headers)
                if h != "" or any(r[ci] != "" for r in data_rows[:100] if ci < len(r))]
        if len(keep) < len(headers):
            headers = [headers[ci] for ci in keep]
            data_rows = [[r[ci] if ci < len(r) else "" for ci in keep] for r in data_rows]

    if not headers and not data_rows:
        return None

    return {
        "title": "分区表",
        "type": "zone_table",
        "headers": headers,
        "rows": data_rows,
    }


def _build_allied_zone_map(wb):
    """Build {postcode: zone} from the 'Allied分区表' sheet in the workbook."""
    if "Allied分区表" not in wb.sheetnames:
        return None
    ws = wb["Allied分区表"]
    pcmap = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 4:
            continue
        pc = str(row[1]).strip() if row[1] else ""
        zone = str(row[3]).strip() if row[3] else ""
        if not pc or not zone:
            continue
        pc = pc.split(".")[0]
        if len(pc) == 3:
            pcmap[pc] = zone
            pcmap["0" + pc] = zone
        elif len(pc) == 4 and pc.isdigit():
            pcmap[pc] = zone
    return pcmap if pcmap else None


def _build_postcode_zone_map(zone_section):
    """Build {postcode: zone_name} from zone_section rows.

    Detects postcode and zone column indices from either the section's headers
    or from a sub-header row within the data.
    """
    if not zone_section:
        return None
    rows = zone_section.get("rows", [])
    headers = zone_section.get("headers", [])
    if not rows:
        return None

    zone_col = None
    postcode_col = None
    data_start = 0

    # First try to find from section headers
    for ci, cell in enumerate(headers):
        s = str(cell).strip().lower()
        if s in ("postcode", "postcodes", "邮编"):
            postcode_col = ci
        if s in ("zone", "price zone", "分区"):
            zone_col = ci

    # If not found in headers, scan rows for sub-header
    if postcode_col is None or zone_col is None:
        postcode_col = None
        zone_col = None
        for ri, row in enumerate(rows):
            for ci, cell in enumerate(row):
                s = str(cell).strip().lower()
                if s in ("postcode", "postcodes", "邮编"):
                    postcode_col = ci
                if s in ("zone", "price zone", "分区"):
                    zone_col = ci
            if postcode_col is not None and zone_col is not None:
                data_start = ri + 1
                break

    if postcode_col is None or zone_col is None:
        return None

    mapping = {}
    for row in rows[data_start:]:
        if len(row) <= max(postcode_col, zone_col):
            continue
        postcode_raw = str(row[postcode_col]).strip()
        zone_raw = str(row[zone_col]).strip()
        if not postcode_raw or not zone_raw:
            continue
        digits = "".join(c for c in postcode_raw if c.isdigit())
        if len(digits) < 3 or len(digits) > 4:
            continue
        zone_name = ZONE_NAME_MAP.get(zone_raw, zone_raw)
        if len(digits) == 3:
            mapping[digits] = zone_name
            mapping["0" + digits] = zone_name
        else:
            mapping[digits] = zone_name

    return mapping if mapping else None


def parse_sheet(ws, sheet_name, key, wb=None):
    """Main dispatcher for parsing a sheet."""
    if key == "mulu":
        sections = parse_mulu(ws)
        return {"name": sheet_name, "key": key, "sections": sections, "is_large": False}

    if key in ("churuku", "zengzhi", "fba"):
        sections = parse_simple_table(ws)
        return {"name": sheet_name, "key": key, "sections": sections, "is_large": False}

    if key == "peichang":
        sections = parse_peichang(ws)
        return {"name": sheet_name, "key": key, "sections": sections, "is_large": False}

    if key == "aupost_letter":
        sections = parse_aupost_letter(ws)
        return {"name": sheet_name, "key": key, "sections": sections, "is_large": False}

    # iMile has a unique layout: zone tables below price tables, 2-row headers
    if key == "imile":
        price_sections, zone_section, postcode_zone_maps = parse_imile(ws, sheet_name)
        sections = []
        for ps in price_sections:
            ps["type"] = "price_table"
            sections.append(ps)
        if zone_section:
            sections.append(zone_section)
        total_rows = sum(len(s.get("rows", [])) for s in sections)
        is_large = total_rows > LARGE_THRESHOLD
        result = {"name": sheet_name, "key": key, "sections": sections, "is_large": is_large}
        if postcode_zone_maps:
            result["postcode_zone_maps"] = postcode_zone_maps
        return result

    # AU POST Eparcel/Express: side-by-side Sydney/Melbourne tables + zone table below
    if key in ("aupost_eparcel", "aupost_express"):
        price_sections, zone_section, postcode_zone_map = parse_aupost_dual(ws, sheet_name)
        sections = []
        for ps in price_sections:
            if ps.get("type") != "richtext":
                ps["type"] = "price_table"
            sections.append(ps)
        if zone_section:
            sections.append(zone_section)
        total_rows = sum(len(s.get("rows", [])) for s in sections)
        is_large = total_rows > LARGE_THRESHOLD
        result = {"name": sheet_name, "key": key, "sections": sections, "is_large": is_large}
        if postcode_zone_map:
            result["postcode_zone_map"] = postcode_zone_map
        return result

    # YTO: stacked price tables + notes + zone table below
    if key == "yto":
        price_sections, zone_section = parse_yto(ws, sheet_name)
        sections = []
        for ps in price_sections:
            if ps.get("type") != "richtext":
                ps["type"] = "price_table"
            sections.append(ps)
        if zone_section:
            sections.append(zone_section)
        postcode_zone_map = _build_postcode_zone_map(zone_section) if zone_section else None
        total_rows = sum(len(s.get("rows", [])) for s in sections)
        result = {"name": sheet_name, "key": key, "sections": sections, "is_large": total_rows > LARGE_THRESHOLD}
        if postcode_zone_map:
            result["postcode_zone_map"] = postcode_zone_map
        return result

    # Fastway: stacked price + zone on right
    if key == "fastway":
        price_sections, zone_section = parse_fastway(ws, sheet_name)
        sections = []
        for ps in price_sections:
            if ps.get("type") != "richtext":
                ps["type"] = "price_table"
            sections.append(ps)
        if zone_section:
            sections.append(zone_section)
        postcode_zone_map = _build_postcode_zone_map(zone_section) if zone_section else None
        total_rows = sum(len(s.get("rows", [])) for s in sections)
        result = {"name": sheet_name, "key": key, "sections": sections, "is_large": total_rows > LARGE_THRESHOLD}
        if postcode_zone_map:
            result["postcode_zone_map"] = postcode_zone_map
        return result

    # Allied: side-by-side, no zone table in this sheet; zone map from allied_zone sheet
    if key == "allied":
        secs, _ = parse_dual_simple(ws, sheet_name, (0, 5), (7, 12))
        sections = []
        for ps in secs:
            if ps.get("type") != "richtext":
                ps["type"] = "price_table"
            sections.append(ps)
        total_rows = sum(len(s.get("rows", [])) for s in sections)
        # Build postcode_zone_map from allied_zone sheet
        postcode_zone_map = _build_allied_zone_map(wb) if wb else None
        result = {"name": sheet_name, "key": key, "sections": sections, "is_large": total_rows > LARGE_THRESHOLD}
        if postcode_zone_map:
            result["postcode_zone_map"] = postcode_zone_map
        return result

    # Border: side-by-side + notes col 12-14 + zone below
    if key == "border":
        secs, zone_section = parse_dual_simple(ws, sheet_name, (0, 4), (6, 10), notes_cols=(11, 15), zone_below=True)
        sections = []
        for ps in secs:
            if ps.get("type") != "richtext":
                ps["type"] = "price_table"
            sections.append(ps)
        if zone_section:
            sections.append(zone_section)
        postcode_zone_map = _build_postcode_zone_map(zone_section) if zone_section else None
        total_rows = sum(len(s.get("rows", [])) for s in sections)
        result = {"name": sheet_name, "key": key, "sections": sections, "is_large": total_rows > LARGE_THRESHOLD}
        if postcode_zone_map:
            result["postcode_zone_map"] = postcode_zone_map
        return result

    # TFM: side-by-side + notes col 12 + zone below
    if key == "tfm":
        secs, zone_section = parse_dual_simple(ws, sheet_name, (0, 4), (6, 10), notes_cols=(11, 15), zone_below=True)
        sections = []
        for ps in secs:
            if ps.get("type") != "richtext":
                ps["type"] = "price_table"
            sections.append(ps)
        if zone_section:
            sections.append(zone_section)
        postcode_zone_map = _build_postcode_zone_map(zone_section) if zone_section else None
        total_rows = sum(len(s.get("rows", [])) for s in sections)
        result = {"name": sheet_name, "key": key, "sections": sections, "is_large": total_rows > LARGE_THRESHOLD}
        if postcode_zone_map:
            result["postcode_zone_map"] = postcode_zone_map
        return result

    # TOLL: side-by-side + notes below + zone on right
    if key == "toll":
        secs, zone_section = parse_toll(ws, sheet_name)
        sections = []
        for ps in secs:
            if ps.get("type") != "richtext":
                ps["type"] = "price_table"
            sections.append(ps)
        if zone_section:
            sections.append(zone_section)
        postcode_zone_map = _build_postcode_zone_map(zone_section) if zone_section else None
        total_rows = sum(len(s.get("rows", [])) for s in sections)
        result = {"name": sheet_name, "key": key, "sections": sections, "is_large": total_rows > LARGE_THRESHOLD}
        if postcode_zone_map:
            result["postcode_zone_map"] = postcode_zone_map
        return result

    has_zone = key == "pfl"
    if has_zone:
        price_sections, zone_section = parse_pricing_with_zones(ws, sheet_name)
    else:
        price_sections, zone_section = parse_generic_large(ws, sheet_name)

    sections = []
    for ps in price_sections:
        ps["type"] = "price_table"
        sections.append(ps)
    if zone_section:
        sections.append(zone_section)

    # Build postcode → zone mapping from zone_section data
    postcode_zone_map = None
    if zone_section:
        postcode_zone_map = _build_postcode_zone_map(zone_section)

    total_rows = sum(len(s.get("rows", [])) for s in sections)
    is_large = total_rows > LARGE_THRESHOLD

    result = {"name": sheet_name, "key": key, "sections": sections, "is_large": is_large}
    if postcode_zone_map:
        result["postcode_zone_map"] = postcode_zone_map
    return result


def main():
    if not os.path.isfile(EXCEL_PATH):
        print(f"ERROR: Excel not found: {EXCEL_PATH}")
        sys.exit(1)

    os.makedirs(OUT_DIR, exist_ok=True)
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    index = []
    for sheet_name in wb.sheetnames:
        key = SHEET_KEY_MAP.get(sheet_name)
        if not key:
            print(f"  SKIP unknown sheet: {sheet_name}")
            continue

        ws = wb[sheet_name]
        print(f"  Parsing: {sheet_name} -> {key}.json ...")
        data = parse_sheet(ws, sheet_name, key, wb=wb)

        total_rows = sum(len(s.get("rows", [])) for s in data.get("sections", []))
        out_path = os.path.join(OUT_DIR, f"{key}.json")
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, separators=(",", ":"))
        size_kb = os.path.getsize(out_path) / 1024
        print(f"    -> {total_rows} rows, {size_kb:.1f} KB")

        index.append({
            "key": key,
            "name": sheet_name,
            "row_count": total_rows,
            "is_large": data.get("is_large", False),
        })

    index_path = os.path.join(OUT_DIR, "_index.json")
    with open(index_path, "w", encoding="utf-8") as f:
        json.dump(index, f, ensure_ascii=False, indent=2)

    print(f"\nDone! {len(index)} sheets -> {OUT_DIR}")
    print(f"Index: {index_path}")


if __name__ == "__main__":
    main()
