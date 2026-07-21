# -*- coding: utf-8 -*-
"""文章详情导出为 xlsx。"""
import io
import json
import math
import os
import re
from urllib.parse import urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

import config


TOTAL_COLUMNS = 10
SHEET_NAME = '报价表'
DEFAULT_ROW_HEIGHT_PT = 15
IMAGE_ROW_HEIGHT_PT = 14
PX_PER_POINT = 4.0 / 3.0
TITLE_FILL = PatternFill('solid', fgColor='2563EB')
TITLE_FONT = Font(name='Microsoft YaHei', size=14, bold=True, color='FFFFFF')
MAIN_TITLE_FONT = Font(name='Microsoft YaHei', size=20, bold=True, color='0F172A')
BODY_FONT = Font(name='Microsoft YaHei', size=12, color='0F172A')
META_FONT = Font(name='Microsoft YaHei', size=10, color='64748B')
THIN_SIDE = Side(style='thin', color='CBD5E1')
HEADER_SIDE = Side(style='medium', color='93C5FD')
TABLE_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
HEADER_BORDER = Border(left=HEADER_SIDE, right=HEADER_SIDE, top=HEADER_SIDE, bottom=HEADER_SIDE)
# 与文章页 .dg-grid-tr-tidan-header 一致（#FFC000）
DG_TIDAN_FILL = PatternFill("solid", fgColor="FFC000")
HEADER_FILL = PatternFill('solid', fgColor='1D4ED8')
HEADER_FONT = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
CELL_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
TEXT_ALIGNMENT = Alignment(horizontal='left', vertical='top', wrap_text=True)
LINK_FONT = Font(name='Microsoft YaHei', size=11, color='0563C1', underline='single')
QUICK_LINKS = [
    ('公司官网', 'https://www.aubvip.com'),
    ('澳邮POST网站', 'https://auspost.com.au/'),
    ('TGE网站', 'https://teamglobalexp.com/'),
    ('TFM网站', 'https://www.tfmxpress.com.au/index.html'),
    ('Hunter网站', 'https://www.hunterexpress.com.au/'),
    ('Allied网站', 'https://www.alliedexpress.com.au/'),
]


def build_article_workbook(article, modules, exported_at_text):
    """生成工作簿。不含页头（文章标题/分类/导出时间/快捷链接），从各模块表格与正文起写。

    article、exported_at_text 仍由 /export-xlsx 传入以保持接口稳定，表中不再写入。
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = None
    sheet.sheet_format.defaultRowHeight = DEFAULT_ROW_HEIGHT_PT

    for idx in range(1, TOTAL_COLUMNS + 1):
        sheet.column_dimensions[get_column_letter(idx)].width = 14

    row = 1
    for module in modules:
        content = _loads_content(module.get('content'))
        module_type = (module.get('type') or '').strip().lower()
        if module_type in {'text', 'richtext'}:
            row = _write_text_module(sheet, row, content)
        elif module_type == 'image':
            row = _write_image_module(sheet, row, content)
        elif module_type == 'channel':
            row = _write_channel_module(sheet, row, content)
        elif module_type == 'dg_grid':
            row = _write_dg_grid_module(sheet, row, content)
        elif module_type in {'surcharge_cn', 'surcharge_intl', 'surcharge'}:
            row = _write_surcharge_module(sheet, row, module_type, content)
        elif module_type == 'overseas_warehouse':
            row = _write_overseas_warehouse_module(sheet, row)
        else:
            row = _write_unknown_module(sheet, row, content)
        row += 1

    return workbook


def workbook_to_bytes(workbook):
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _write_main_title(sheet, row, title):
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLUMNS)
    cell = sheet.cell(row=row, column=1, value=title)
    cell.font = MAIN_TITLE_FONT
    cell.alignment = Alignment(horizontal='left', vertical='center')
    sheet.row_dimensions[row].height = 28
    return row + 1


def _write_section_title(sheet, row, title):
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLUMNS)
    cell = sheet.cell(row=row, column=1, value=title)
    cell.fill = TITLE_FILL
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal='left', vertical='center')
    sheet.row_dimensions[row].height = 24
    return row + 1


def _write_merged_text(sheet, row, text, font=None, alignment=None, fill=None, height=None):
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLUMNS)
    cell = sheet.cell(row=row, column=1, value=text)
    cell.font = font or BODY_FONT
    cell.alignment = alignment or TEXT_ALIGNMENT
    if fill is not None:
        cell.fill = fill
    if height is None:
        height = _estimate_text_height(text, TOTAL_COLUMNS)
    sheet.row_dimensions[row].height = height
    return row + 1


def _write_text_module(sheet, row, content):
    title = (content.get('title') or '').strip()
    body_html = content.get('html') if content.get('html') is not None else content.get('text')
    remark_html = content.get('remark')

    if title:
        row = _write_section_title(sheet, row, title)
    for block in _html_to_blocks(body_html):
        row = _write_merged_text(sheet, row, block, font=BODY_FONT, alignment=TEXT_ALIGNMENT)
    if remark_html:
        for block in _html_to_blocks(remark_html):
            row = _write_merged_text(sheet, row, block, font=BODY_FONT, alignment=TEXT_ALIGNMENT)
    return row


def _write_image_module(sheet, row, content):
    image_url = content.get('url')
    caption = (content.get('caption') or '').strip()
    remark_html = content.get('remark')

    if image_url:
        image_bytes = _read_image_bytes(image_url)
        if image_bytes is not None:
            row = _insert_image(sheet, row, image_bytes)
    if caption:
        row = _write_merged_text(
            sheet, row, caption, font=Font(name='Microsoft YaHei', size=11, italic=True, color='475569'),
            alignment=Alignment(horizontal='center', vertical='center', wrap_text=True), height=18
        )
    if remark_html:
        for block in _html_to_blocks(remark_html):
            row = _write_merged_text(sheet, row, block, font=BODY_FONT, alignment=TEXT_ALIGNMENT)
    return row


def _write_channel_module(sheet, row, content):
    title = (content.get('title') or '').strip()
    if title:
        row = _write_section_title(sheet, row, title)

    weights = list(content.get('weights') or [])
    states = list(content.get('states') or [])
    headers = ['目的港', '邮编', '区域'] + [str(w) for w in weights] + ['到门时效', '渠道']

    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=col_index, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = HEADER_BORDER
        cell.alignment = CELL_ALIGNMENT
    sheet.row_dimensions[row].height = 22
    header_row = row
    row += 1

    data_start_row = row
    for state in states:
        values = [
            state.get('name_cn', ''),
            str(state.get('postcode_range', '') or ''),
            state.get('area_type') or '-',
        ]
        prices = list(state.get('prices') or [])
        for idx in range(len(weights)):
            values.append(prices[idx] if idx < len(prices) and prices[idx] else '—')
        values.extend([state.get('delivery_time', ''), state.get('channel', '')])

        for col_index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=col_index, value=str(value) if value is not None else '')
            cell.border = TABLE_BORDER
            cell.font = BODY_FONT
            cell.alignment = _alignment_for_header(headers[col_index - 1])
        sheet.row_dimensions[row].height = _estimate_row_height(values)
        row += 1

    if row > data_start_row:
        _merge_table_columns(sheet, data_start_row, row - 1, headers)

    remark_html = content.get('remark')
    if remark_html:
        row += 1
        for block in _html_to_blocks(remark_html):
            row = _write_merged_text(sheet, row, block, font=BODY_FONT, alignment=TEXT_ALIGNMENT)
    else:
        row += 1

    _autosize_channel_columns(sheet, header_row, row - 1, headers)
    return row


def _write_dg_grid_v2_module(sheet, row, c):
    """DG v2：抬头、客户、单证、费用行、备注（简要导出）。"""
    from openpyxl.styles import Alignment, Font

    hm, hs = c.get("header_main") or "", c.get("header_sub") or ""
    if hm or hs:
        row = _write_merged_text(
            sheet,
            row,
            (hm or "") + (" · " if hm and hs else "") + (hs or ""),
            font=Font(name="Microsoft YaHei", size=15, bold=True, color="0F172A"),
            alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
            height=26,
        )
    for rowobj in c.get("customer_rows") or []:
        lab, val = (rowobj.get("label") or ""), (rowobj.get("value") or "")
        if not str(lab).strip() and not str(val).strip():
            continue
        row = _write_merged_text(
            sheet,
            row,
            f"{lab} {val}".strip(),
            font=Font(name="Microsoft YaHei", size=11, color="0F172A"),
            alignment=TEXT_ALIGNMENT,
            height=None,
        )
    log = c.get("logistics") or {}
    log_pairs = [
        ("起运港", "origin_port"),
        ("中转港", "transit_port"),
        ("目的港", "dest_port"),
        ("海外仓", "dest_warehouse"),
        ("船名航次", "vessel"),
    ]
    log_parts = [f"{lab}：{log.get(k) or ''}" for lab, k in log_pairs if (log.get(k) or "").strip()]
    if log_parts:
        row = _write_merged_text(
            sheet,
            row,
            "　".join(log_parts),
            font=BODY_FONT,
            alignment=TEXT_ALIGNMENT,
        )
    row = _write_section_title(sheet, row, "费用项目")
    dual = bool(c.get("fee_dual_amount"))
    if dual:
        h1 = str(c.get("fee_amount_label_1") or "20GP/金额")
        h2 = str(c.get("fee_amount_label_2") or "40HQ/金额")
        headers = ["提单号", "柜号", "费用项目", "币别", h1, h2, "说明"]
        ncol = 7
    else:
        headers = ["提单号", "柜号", "费用项目", "币别", "金额", "说明"]
        ncol = 6
    for i, h in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = HEADER_BORDER
        cell.alignment = CELL_ALIGNMENT
    sheet.row_dimensions[row].height = 20
    row += 1
    for it in c.get("fee_items") or []:
        is_sub = (it.get("row_kind") or "item") == "subtotal"
        if is_sub:
            for j, v in enumerate(
                [it.get("bl_no", ""), it.get("container_no", "")], start=1
            ):
                cell = sheet.cell(row=row, column=j, value=str(v) if v is not None else "")
                cell.font = BODY_FONT
                cell.border = TABLE_BORDER
                cell.alignment = TEXT_ALIGNMENT
            sheet.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
            cname = sheet.cell(
                row=row, column=3, value=str(it.get("name") or "")
            )
            cname.font = Font(name="Microsoft YaHei", size=12, color="0F172A", bold=True)
            cname.border = TABLE_BORDER
            cname.alignment = TEXT_ALIGNMENT
            cols_write = [(5, it.get("amount", "")), (6, it.get("note", ""))]
            if dual:
                cols_write = [
                    (5, it.get("amount", "")),
                    (6, it.get("amount_alt", "")),
                    (7, it.get("note", "")),
                ]
            for j, v in cols_write:
                cell = sheet.cell(
                    row=row, column=j, value=str(v) if v is not None else ""
                )
                cell.font = BODY_FONT
                cell.border = TABLE_BORDER
                cell.alignment = TEXT_ALIGNMENT
            line = [
                it.get("bl_no", ""),
                it.get("container_no", ""),
                it.get("name", ""),
                it.get("amount", ""),
            ]
            if dual:
                line.extend([it.get("amount_alt", ""), it.get("note", "")])
            else:
                line.append(it.get("note", ""))
        else:
            line = [
                it.get("bl_no", ""),
                it.get("container_no", ""),
                it.get("name", ""),
                it.get("currency", ""),
                it.get("amount", ""),
            ]
            if dual:
                line.extend([it.get("amount_alt", ""), it.get("note", "")])
            else:
                line.append(it.get("note", ""))
            for j, v in enumerate(line, start=1):
                cell = sheet.cell(row=row, column=j, value=str(v) if v is not None else "")
                cell.font = BODY_FONT
                cell.border = TABLE_BORDER
                cell.alignment = TEXT_ALIGNMENT
        sheet.row_dimensions[row].height = max(
            16, _estimate_row_height([str(x) for x in line])
        )
        row += 1
    for cc in range(1, ncol + 1):
        sheet.column_dimensions[get_column_letter(cc)].width = min(20, 10 + (cc * 0.3))
    remark = c.get("remark")
    if remark:
        for block in _html_to_blocks(remark):
            row = _write_merged_text(sheet, row, block, font=BODY_FONT, alignment=TEXT_ALIGNMENT)
    return row


def _write_dg_grid_module(sheet, row, content):
    """v2 为分块报价；旧数据为合并矩阵表。"""
    from services.dg_quote_grid import is_dg_content_v2, merge_dg_v2_content

    c = content if isinstance(content, dict) else _loads_content(content)
    c = c if isinstance(c, dict) else {}
    if is_dg_content_v2(c):
        return _write_dg_grid_v2_module(sheet, row, merge_dg_v2_content(c))
    from openpyxl.styles import Alignment, Font

    from services.dg_quote_grid import (
        _prepare_dg_table_structure,
        canonical_excel_grid_module_content,
        prepare_dg_table_display,
    )

    c = canonical_excel_grid_module_content(dict(c))
    tid = ((c or {}).get("template_id") or "").strip()
    if tid == "普柜":
        fb = "普柜报价表"
    elif tid == "9类电池柜":
        fb = "9类电池柜报价表"
    else:
        fb = "报价表"
    title = (c or {}).get("title") or fb
    row = _write_merged_text(
        sheet,
        row,
        title,
        font=Font(name="Microsoft YaHei", size=16, bold=True, color="0F172A"),
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
        height=28,
    )

    cells, merges, hr_in = prepare_dg_table_display(c or {})
    cells, merges, n, m, omit, header_orange_row = _prepare_dg_table_structure(
        cells, merges, hr_in
    )
    base_row = row
    dg_font = Font(name="Microsoft YaHei", size=11, color="0F172A")
    dg_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r in range(n):
        for cc in range(m):
            op = omit[r][cc]
            if op is True:
                continue
            val = cells[r][cc] if cc < len(cells[r]) else ""
            st = str(val) if val is not None else ""
            cell = sheet.cell(row=base_row + r, column=1 + cc, value=st)
            cell.font = dg_font
            cell.alignment = dg_align
            cell.border = TABLE_BORDER
            if header_orange_row is not None and r == header_orange_row:
                cell.fill = DG_TIDAN_FILL
        visible = [
            str(cells[r][cc] or "")
            for cc in range(m)
            if omit[r][cc] is not True
        ]
        sheet.row_dimensions[base_row + r].height = _estimate_row_height(visible)

    for mg in merges:
        r, c, rs, cs = int(mg["r"]), int(mg["c"]), int(mg["rs"]), int(mg["cs"])
        if r < 0 or c < 0 or r + rs > n or c + cs > m:
            continue
        if rs == 1 and cs == 1:
            continue
        sheet.merge_cells(
            start_row=base_row + r,
            start_column=1 + c,
            end_row=base_row + r + rs - 1,
            end_column=1 + c + cs - 1,
        )

    for cc in range(min(m, TOTAL_COLUMNS)):
        sheet.column_dimensions[get_column_letter(1 + cc)].width = 11.5

    row = base_row + n
    remark = (c or {}).get("remark")
    if remark:
        for block in _html_to_blocks(remark):
            row = _write_merged_text(
                sheet, row, block, font=BODY_FONT, alignment=TEXT_ALIGNMENT
            )
    return row


def _write_unknown_module(sheet, row, content):
    text = json.dumps(content, ensure_ascii=False, indent=2) if content else ''
    return _write_merged_text(sheet, row, text, font=BODY_FONT, alignment=TEXT_ALIGNMENT)


def _write_surcharge_module(sheet, row, module_type, content=None):
    """附加明细：导出为提示行（详情见网页版）。标题优先取用户编辑的 content.title。"""
    title = ''
    if isinstance(content, dict):
        title = (content.get('title') or '').strip()
    if not title:
        title = '国外附加明细' if module_type == 'surcharge_intl' else '国内附加明细'
    row = _write_section_title(sheet, row, title)
    return _write_merged_text(
        sheet, row,
        f'{title}为收费标准表，请在网页版文章中查看完整表格。',
        font=BODY_FONT, alignment=TEXT_ALIGNMENT,
    )


def _write_overseas_warehouse_module(sheet, row):
    """常见海外仓地址：导出为提示行（详情见网页版）。"""
    row = _write_section_title(sheet, row, '常见海外仓地址')
    return _write_merged_text(
        sheet, row,
        '常见海外仓地址明细请在网页版文章中查看完整表格。',
        font=BODY_FONT, alignment=TEXT_ALIGNMENT,
    )


def _write_quick_links(sheet, row):
    row = _write_section_title(sheet, row, '快捷链接')
    for label, url in QUICK_LINKS:
        row = _write_link_row(sheet, row, label, url)
    return row


def _write_link_row(sheet, row, label, url):
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLUMNS)
    cell = sheet.cell(row=row, column=1, value=label)
    cell.hyperlink = url
    cell.font = LINK_FONT
    cell.alignment = Alignment(horizontal='left', vertical='center')
    sheet.row_dimensions[row].height = 18
    return row + 1


def _insert_image(sheet, row, image_bytes):
    image_stream = io.BytesIO(image_bytes)
    with PILImage.open(image_stream) as pil_image:
        original_width, original_height = pil_image.size
    image_stream.seek(0)
    xl_image = XLImage(image_stream)
    max_width_px = 980
    if original_width > max_width_px:
        scale = max_width_px / float(original_width)
    else:
        scale = 1.0
    xl_image.width = int(original_width * scale)
    xl_image.height = int(original_height * scale)
    anchor = 'A{row}'.format(row=row)
    sheet.add_image(xl_image, anchor)

    row_height_px = IMAGE_ROW_HEIGHT_PT * PX_PER_POINT
    occupied_rows = max(1, int(math.ceil(xl_image.height / row_height_px)))
    for current_row in range(row, row + occupied_rows):
        sheet.row_dimensions[current_row].height = IMAGE_ROW_HEIGHT_PT
    return row + occupied_rows


def _merge_table_columns(sheet, start_row, end_row, headers):
    merge_indices = []
    for idx, header in enumerate(headers, start=1):
        if header in {'目的港', '邮编', '区域', '到门时效', '渠道'}:
            merge_indices.append(idx)
            continue
        if any(key in header for key in ('备注', '澳洲小包限定标准', '挂号费', '单价', '首', '续', '重量')) or 'kg' in header.lower():
            merge_indices.append(idx)

    for col_index in merge_indices:
        current_start = start_row
        current_value = str(sheet.cell(row=start_row, column=col_index).value or '')
        for row in range(start_row + 1, end_row + 2):
            value = str(sheet.cell(row=row, column=col_index).value or '') if row <= end_row else None
            if row <= end_row and value == current_value and current_value not in {'', '—'}:
                continue
            if row - current_start > 1 and current_value not in {'', '—'}:
                sheet.merge_cells(start_row=current_start, start_column=col_index, end_row=row - 1, end_column=col_index)
                merged_cell = sheet.cell(row=current_start, column=col_index)
                merged_cell.alignment = _alignment_for_header(headers[col_index - 1])
            current_start = row
            current_value = value


def _autosize_channel_columns(sheet, start_row, end_row, headers):
    for idx, header in enumerate(headers, start=1):
        width = 12
        if '备注' in header:
            width = 22
        elif '澳洲小包限定标准' in header:
            width = 28
        elif header in {'邮编', '渠道', '到门时效'}:
            width = 14
        elif header in {'目的港', '区域'}:
            width = 11
        elif any(key in header for key in ('挂号费', '单价', '首', '续', '重量')) or 'kg' in header.lower():
            width = 11
        sheet.column_dimensions[get_column_letter(idx)].width = width


def _alignment_for_header(header):
    if '备注' in header or '标准' in header:
        return TEXT_ALIGNMENT
    return CELL_ALIGNMENT


def _estimate_text_height(text, width_columns):
    normalized = str(text or '').strip()
    if not normalized:
        return 18
    line_count = 0
    for line in normalized.splitlines() or ['']:
        visual_len = max(1, len(line))
        line_count += max(1, int(math.ceil(visual_len / max(1, width_columns * 4))))
    return max(18, min(180, 17 * line_count))


def _estimate_row_height(values):
    longest = max((len(str(value or '')) for value in values), default=1)
    return max(18, min(54, 16 * max(1, int(math.ceil(longest / 20.0)))))


def _loads_content(raw):
    if not raw:
        return {}
    if isinstance(raw, dict):
        return raw
    try:
        return json.loads(raw)
    except Exception:
        return {}


def _html_to_blocks(html):
    if not html:
        return []
    soup = BeautifulSoup(str(html), 'html.parser')
    blocks = []
    for node in soup.find_all(['h1', 'h2', 'h3', 'h4', 'p', 'li', 'blockquote']):
        text = node.get_text(' ', strip=True)
        if node.name == 'li':
            text = '• ' + text
        if text:
            blocks.append(text)
    for img in soup.find_all('img'):
        alt = (img.get('alt') or '').strip()
        if alt:
            blocks.append('[图片说明] ' + alt)
    if not blocks:
        text = soup.get_text('\n', strip=True)
        blocks = [line.strip() for line in text.splitlines() if line.strip()]
    return blocks


def _read_image_bytes(url):
    if not url:
        return None

    local_bytes = _read_local_image_bytes(url)
    if local_bytes is not None:
        return local_bytes

    try:
        response = requests.get(url, timeout=15)
        response.raise_for_status()
        return response.content
    except Exception:
        return None


def _read_local_image_bytes(url):
    parsed = urlparse(url)
    path = parsed.path or ''
    if not path.startswith('/'):
        return None
    if path.startswith('/uploads/'):
        file_path = os.path.join(config.UPLOAD_FOLDER, os.path.basename(path))
    elif path.startswith('/static/'):
        relative_path = path[len('/static/'):]
        file_path = os.path.join(config._BASE_DIR, 'static', relative_path)
    else:
        return None
    if not os.path.isfile(file_path):
        return None
    with open(file_path, 'rb') as file_obj:
        return file_obj.read()
