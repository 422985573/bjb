# -*- coding: utf-8 -*-
"""dg_grid 模块：DG报价(schema v2) 与 Excel栅格（9类电池柜报价表、普柜报价表等）分模板维护；见 templates/admin/dg/ 与 templates/partials/。"""
from datetime import date, datetime
import copy
import html
import os
import re

import config

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None  # type: ignore

# 提单号表头行：与 xlsx 中「提单号」行一致，用于全行橙色底
_TIDAN_MARKER = '提单号'
# 仅当「最宽行」列数仍像原始全表（≥13）时才裁，避免对已存好的 9 列数据重复裁切
DG_EDGE_TRIM_MIN_COLS = 13
_DROP_HEAD_COLS = 1
_DROP_TAIL_COLS = 3
_COL_KEEP_LO = 2
_COL_KEEP_HI = 10  # 保留 B–J 共 9 列（1-based）

# 后台编辑：整行合并单元格内「报价日期： 2026-05-13」拆成标签 / 值两个 input
_DG_LV_SPLIT_RE = re.compile(r"^(.+?[：:])(\s*)(.*)$", re.DOTALL)
# DG v2 费用表中 「PS:」「PS：」「P.S.」等免责说明（行首或空白后，避免误伤「DPS」等）
_DG_V2_PS_SNIPPET_RE = re.compile(r"(?:^|[\s\u3000\n])(?:P\s*\.?\s*S\s*\.?\s*[：:])", re.I)


def normalize_dg_v2_ps_rows_to_remark(content):
    """
    将费用表中典型的免责/备注行（以 PS: 开头或含海关查验+实报实销）移入 remark，并从 fee_items 删除。
    幂等：已拆过后不会再匹配。
    """
    if not isinstance(content, dict):
        return content
    try:
        if int(content.get("schema") or 0) != 2:
            return content
    except (TypeError, ValueError):
        return content
    fee = content.get("fee_items")
    if not isinstance(fee, list):
        return content

    def _should_move(it):
        if not isinstance(it, dict):
            return False
        name = str(it.get("name") or "").strip()
        note = str(it.get("note") or "").strip()
        blob = name + (" " + note if note else "")
        if _DG_V2_PS_SNIPPET_RE.search(name) or _DG_V2_PS_SNIPPET_RE.search(note):
            return True
        if "海关查验" in blob and "实报实销" in blob:
            return True
        return False

    def _row_text(it):
        name = str(it.get("name") or "").strip()
        note = str(it.get("note") or "").strip()
        if name and note:
            return name + "\n" + note
        return name or note

    keep = []
    chunks = []
    for it in fee:
        if _should_move(it):
            t = _row_text(it)
            if t:
                chunks.append(t)
        else:
            keep.append(it)
    if not chunks:
        return content

    out = dict(content)
    out["fee_items"] = keep
    r0 = str(out.get("remark") or "").strip()
    deduped = []
    for ch in chunks:
        s = (ch or "").strip()
        if not s:
            continue
        if s in r0 or any(s in exist for exist in deduped):
            continue
        deduped.append(ch.strip())
    tail = "\n\n".join(deduped)
    if tail:
        out["remark"] = (r0 + "\n\n" + tail).strip() if r0 else tail
    else:
        out["remark"] = r0
    return out


def normalize_dg_grid_cell_display_spaces(val):
    """
    前台只读 / 编辑展示：去掉首尾空白，将连续空格（含全角空格、Tab）压成单个半角空格；
    多行时对逐行同样处理。与 DG v2 元数据表观感一致，去掉 xlsx 带入的多余空白。
    """
    if val is None:
        return ""
    s = str(val)
    if "\n" in s:
        lines = []
        for ln in s.split("\n"):
            lines.append(re.sub(r"[\t\u3000 ]{2,}", " ", ln.strip()))
        return "\n".join(lines).strip()
    return re.sub(r"[\t\u3000 ]{2,}", " ", s.strip())


def _abs_path():
    p = config.DG_QUOTE_XLSX_PATH
    return p if os.path.isabs(p) else os.path.join(config._BASE_DIR, p)


def _cell_to_str(value):
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d %H:%M')[: 16] if value.hour or value.minute else value.strftime('%Y-%m-%d')
    if isinstance(value, date):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, float) and value == int(value):
            return str(int(value))
        return str(value)
    return str(value)


def _read_sheet_grid_with_sources(ws):
    max_row, max_col = (ws.max_row or 0), (ws.max_column or 0)
    if max_row < 1 or max_col < 1:
        return 'Sheet1', 1, 1, [['']], [1]

    cells = []
    row_sources = []
    for r in range(1, max_row + 1):
        line = []
        for c in range(1, max_col + 1):
            line.append(_cell_to_str(ws.cell(r, c).value))
        cells.append(line)
        row_sources.append(r)
    return (ws.title or 'Sheet1'), max_row, max_col, cells, row_sources


def _row_is_empty(row):
    if not row:
        return True
    for c in row:
        s = (str(c).strip() if c is not None else '')
        if s:
            return False
    return True


def trim_leading_empty_rows(cells):
    """去掉顶部连续全空行（xlsx 常有一行空表头行）。"""
    if not cells:
        return cells
    out = list(cells)
    while out and _row_is_empty(out[0]):
        out.pop(0)
    return out if out else [['']]


def trim_leading_empty_rows_paired(cells, row_sources):
    """与 trim_leading_empty_rows 同步裁剪 row_sources（1-based Excel 行号）。"""
    if not cells or not row_sources or len(cells) != len(row_sources):
        return cells, row_sources
    out = list(cells)
    out_s = list(row_sources)
    while out and _row_is_empty(out[0]):
        out.pop(0)
        out_s.pop(0)
    if not out:
        return [['']], (out_s[:1] if out_s else [1])
    return out, out_s


def trim_dg_edge_columns(cells):
    """去掉每行左侧第 1 列与右侧最后 3 列（宽表用）。"""
    if not cells:
        return cells
    max_w = max((len(r) for r in cells if r is not None), default=0)
    if max_w < DG_EDGE_TRIM_MIN_COLS:
        return list(cells)
    need = _DROP_HEAD_COLS + _DROP_TAIL_COLS
    out = []
    for row in cells:
        if not row:
            out.append([])
            continue
        r = list(row)
        n = len(r)
        if n <= need:
            out.append([])
        else:
            out.append(r[_DROP_HEAD_COLS : n - _DROP_TAIL_COLS])
    return out if out else [['']]


def normalize_dg_grid_cells(cells):
    """展示/保存/导入用：去顶空行 + 去左 1 列与右 3 列。"""
    if not cells:
        return [['']]
    cells = trim_leading_empty_rows(cells)
    cells = trim_dg_edge_columns(cells)
    return cells if cells else [['']]


def find_tidan_header_row(cells):
    """返回含「提单号」的行的 0-based 索引；找不到返回 None。"""
    if not cells:
        return None
    for i, row in enumerate(cells):
        for c in row or []:
            if _TIDAN_MARKER in str(c or ''):
                return i
    return None


def find_excel_quote_fee_header_row(cells):
    """
    Excel 栅格前台分段：首个「费用明细列标题」行（柜类报价常为第二列表头行）。
    命中条件：同行单元格中同时出现「费用项目」与「币别」/「币种」，或任意单元格含「提单号」。
    """
    if not cells:
        return None
    scan = min(len(cells), 48)
    currency_markers = {'币别', '币种'}
    for i in range(scan):
        row = cells[i] or []
        labels = [str(x or '').strip() for x in row]
        flat = ''.join(labels)
        if _TIDAN_MARKER in flat:
            return i
        if '费用项目' in labels and any(m in labels for m in currency_markers):
            return i
    return None


def _excel_quote_primary_cell_text(row):
    """合并单元格逻辑行：取该行首个非空单元格文本。"""
    if not row:
        return ''
    for x in row:
        t = str(x or '').strip()
        if t:
            return normalize_dg_grid_cell_display_spaces(str(x or ''))
    return ''


def _excel_quote_split_main_sub_title(line):
    """「中海豹国际 · 报价单」→ (main, sub)；无法拆分则 (整行, '')。"""
    s = (line or '').strip()
    if not s:
        return ('', '')
    for sep in (' · ', '·', ' — ', ' - ', '－'):
        if sep in s:
            a, b = s.split(sep, 1)
            a, b = a.strip(), b.strip()
            if a and b:
                return (a, b)
    return (s, '')


def _excel_quote_meta_banner_trim(meta_end, cells):
    """去掉紧贴明细表上方的单列「费用项目」横幅行（改用版块标题展示）。"""
    while meta_end > 0:
        t = _excel_quote_primary_cell_text(cells[meta_end - 1]).strip()
        if t in ('费用项目', '费用明细', '💰 费用项目'):
            meta_end -= 1
            continue
        break
    return meta_end


def _fee_slice_merges(merges, fee_row):
    """仅保留 fee_row 及以下的合并区，行号改为相对 fee_row。"""
    out = []
    for mg in merges or []:
        r = int(mg.get('r', 0))
        if r < fee_row:
            continue
        out.append({
            'r': r - fee_row,
            'c': int(mg.get('c', 0)),
            'rs': int(mg.get('rs', 1)),
            'cs': int(mg.get('cs', 1)),
        })
    return out


def _fee_relative_header_orange_row(full_hr, fee_row, fee_cells):
    """原文 header_orange_row 转为费用子表行号；无效则用 fee_cells 推断。"""
    tid = find_tidan_header_row(fee_cells)
    if full_hr is None or full_hr == '':
        return tid
    try:
        fr = int(full_hr)
    except (TypeError, ValueError):
        return tid
    if fr < fee_row:
        return tid
    return fr - fee_row


def _dg_remove_rows_slice(cells, merges, row_lo, row_hi_exclusive):
    """删除 [row_lo, row_hi_exclusive) 行（0-based），下移 merges；与删除带相交的合并一律丢弃（横幅区通常无合并）。"""
    if (
        not cells
        or row_lo >= row_hi_exclusive
        or row_lo < 0
        or row_hi_exclusive > len(cells)
    ):
        return cells, merges or []
    n_del = row_hi_exclusive - row_lo
    new_cells = cells[:row_lo] + cells[row_hi_exclusive:]
    new_merges = []
    for mg in merges or []:
        try:
            r0 = int(mg["r"])
            rs = int(mg["rs"])
            c0 = int(mg["c"])
            cs = int(mg["cs"])
        except (KeyError, TypeError, ValueError):
            continue
        r1 = r0 + rs - 1
        if r1 < row_lo:
            new_merges.append({"r": r0, "c": c0, "rs": rs, "cs": cs})
        elif r0 >= row_hi_exclusive:
            new_merges.append({"r": r0 - n_del, "c": c0, "rs": rs, "cs": cs})
        else:
            continue
    return new_cells, new_merges


def _dg_row_all_cells_blank_for_strip(cells, r, m):
    """整行去掉空白后是否无任何字符（用于删掉费用表分组之间的占位空行）。"""
    row = cells[r] if 0 <= r < len(cells) else []
    for j in range(m):
        v = row[j] if j < len(row) else ""
        if normalize_dg_grid_cell_display_spaces(str(v or "")).strip():
            return False
    return True


def strip_interior_fully_empty_fee_rows(cells, merges, header_orange_row=None):
    """
    删除费用明细区内「整行无内容」的行（不含抬头表头行及其上方）。
    Excel 模板常在分组之间留空行，div 布局会渲成独立白条；去掉后与表格观感一致。
    """
    cells = [list(r) for r in (cells or [])] if cells else []
    if len(cells) <= 1:
        return cells, list(merges or []), header_orange_row
    fee_hdr = find_excel_quote_fee_header_row(cells)
    min_row = (fee_hdr + 1) if fee_hdr is not None else 1
    m = max((len(r) for r in cells), default=1)
    merges = list(merges or [])
    hr = header_orange_row
    changed = True
    while changed:
        changed = False
        for rdel in range(len(cells) - 1, min_row - 1, -1):
            if not _dg_row_all_cells_blank_for_strip(cells, rdel, m):
                continue
            cells, merges = _dg_remove_rows_slice(cells, merges, rdel, rdel + 1)
            if hr is not None and hr != "":
                try:
                    hi = int(hr)
                    if hi == rdel:
                        hr = None
                    elif hi > rdel:
                        hr = hi - 1
                except (TypeError, ValueError):
                    pass
            changed = True
            break
    return cells, merges, hr


def strip_excel_quote_banner_gap_rows(cells, merges):
    """
    去掉明细表头行之上、仅作横幅用的「费用项目」等占位行（非表格正文）。
    与前台分段逻辑一致，保存与展示共用。
    """
    if not cells:
        return cells or [], merges or []
    fee_hdr = find_excel_quote_fee_header_row(cells)
    if fee_hdr is None:
        return cells, merges or []
    meta_end = _excel_quote_meta_banner_trim(fee_hdr, cells)
    if meta_end >= fee_hdr:
        return cells, merges or []
    return _dg_remove_rows_slice(cells, merges or [], meta_end, fee_hdr)


def _excel_quote_row_non_empty_parts(row):
    parts = []
    for x in row or []:
        t = normalize_dg_grid_cell_display_spaces(str(x or '')).strip()
        if t:
            parts.append(t)
    return parts


def strip_excel_quote_ps_disclaimer_rows(cells, merges, remark):
    """
    Excel 栅格（非 v2）：将表格内 PS / 海关查验+实报实销类免责行移入 remark，并从网格删除。
    规则对齐 normalize_dg_v2_ps_rows_to_remark；幂等。
    """
    if not cells:
        return cells or [], merges or [], str(remark or '').strip()
    merges = list(merges or [])
    cells = [list(r) for r in cells]

    def _should_move(row):
        parts = _excel_quote_row_non_empty_parts(row)
        if not parts:
            return False
        blob = ' '.join(parts)
        if _DG_V2_PS_SNIPPET_RE.search(blob):
            return True
        if '海关查验' in blob and '实报实销' in blob:
            return True
        return False

    def _row_remark_text(row):
        parts = _excel_quote_row_non_empty_parts(row)
        if len(parts) >= 2:
            return '\n'.join(parts)
        return parts[0] if parts else ''

    indices = [i for i, row in enumerate(cells) if _should_move(row)]
    if not indices:
        return cells, merges, str(remark or '').strip()

    chunks = []
    for i in indices:
        t = _row_remark_text(cells[i])
        if t:
            chunks.append(t)

    r0 = str(remark or '').strip()
    deduped = []
    for ch in chunks:
        s = (ch or '').strip()
        if not s:
            continue
        if s in r0 or any(s in exist for exist in deduped):
            continue
        deduped.append(s)

    for i in sorted(indices, reverse=True):
        cells, merges = _dg_remove_rows_slice(cells, merges, i, i + 1)

    tail = '\n\n'.join(deduped)
    if tail:
        new_r = (r0 + '\n\n' + tail).strip() if r0 else tail
    else:
        new_r = r0
    return cells, merges, new_r


def render_dg_excel_grid_split_document_html(cells, merges, header_orange_row_full, fee_row):
    """
    前台 Excel 栅格：抬头区用 dg-quote-v2-meta 四列表单（对齐 DG v2 观感），费用明细仍为表格。
    """
    from markupsafe import Markup, escape

    meta_end = _excel_quote_meta_banner_trim(fee_row, cells)
    meta_slice = cells[:meta_end] if meta_end > 0 else []
    fee_cells = cells[fee_row:] if fee_row < len(cells) else []
    fee_merges = _fee_slice_merges(merges, fee_row)
    fee_hr = _fee_relative_header_orange_row(header_orange_row_full, fee_row, fee_cells)

    parts = ['<div class="dg-quote-v2 dg-quote-v2--grid dg-excel-quote-layout">']

    parts.append(
        '<table class="dg-quote-v2-meta dg-quote-v2-meta--excel" '
        'role="presentation" aria-label="报价抬头与概要">'
        '<colgroup>'
        '<col style="width:18%"><col style="width:32%">'
        '<col style="width:18%"><col style="width:32%">'
        '</colgroup>'
    )
    idx = 0
    if meta_slice:
        line0 = _excel_quote_primary_cell_text(meta_slice[0])
        main, sub = _excel_quote_split_main_sub_title(line0)
        looks_kv = split_dg_editable_label_value(
            normalize_dg_grid_cell_display_spaces(line0)
        )
        if main and not looks_kv:
            if main:
                parts.append(
                    f'<tr class="dg-c-main"><th colspan="4" scope="col">{escape(main)}</th></tr>'
                )
            if sub:
                parts.append(
                    f'<tr class="dg-c-sub"><th colspan="4" scope="col">{escape(sub)}</th></tr>'
                )
            idx = 1

    ordered_meta = []  # 保留文档顺序：(kv,) / plain
    for r in range(idx, len(meta_slice)):
        text = _excel_quote_primary_cell_text(meta_slice[r])
        if not str(text or '').strip():
            continue
        cleaned = normalize_dg_grid_cell_display_spaces(text)
        lv = split_dg_editable_label_value(cleaned)
        if lv:
            lab, gap, val = lv
            ordered_meta.append(
                ('kv', lab + (gap or ''), val),
            )
        else:
            ordered_meta.append(('plain', cleaned))

    pair_row_buf = []

    def _flush_kv_pairs_two_even(buf):
        i = 0
        while i < len(buf):
            a = buf[i]
            b = buf[i + 1] if i + 1 < len(buf) else None
            parts.append('<tr class="dg-c-kv">')
            parts.append(f'<th scope="row" class="dg-c-lab">{escape(str(a[0]))}</th>')
            if b is not None:
                parts.append(f'<td class="dg-c-val">{escape(str(a[1]))}</td>')
                parts.append(f'<th scope="row" class="dg-c-lab">{escape(str(b[0]))}</th>')
                parts.append(f'<td class="dg-c-val">{escape(str(b[1]))}</td></tr>')
            else:
                # 奇数个键值对：值格 colspan 占满右侧，避免空白「假列」
                parts.append(
                    f'<td class="dg-c-val dg-c-val--wide" colspan="3">{escape(str(a[1]))}</td></tr>'
                )
            i += 2

    for tag in ordered_meta:
        if tag[0] == 'plain':
            _flush_kv_pairs_two_even(pair_row_buf)
            pair_row_buf.clear()
            parts.append(
                f'<tr class="dg-c-notes"><td colspan="4">{escape(tag[1])}</td></tr>'
            )
        else:
            pair_row_buf.append((tag[1], tag[2]))
            if len(pair_row_buf) >= 2:
                _flush_kv_pairs_two_even(pair_row_buf)
                pair_row_buf.clear()
    _flush_kv_pairs_two_even(pair_row_buf)
    parts.append('</table>')

    parts.append('<div class="dg-quote-v2-section dg-quote-v2-fee-block">')
    parts.append('<h3 class="dg-quote-v2-sec-title">费用项目</h3>')
    parts.append('<div class="dg-quote-v2-fee-wrap">')
    parts.append(render_dg_table_html(fee_cells, fee_merges, fee_hr))
    parts.append('</div></div></div>')
    return Markup(''.join(parts))


def find_shuoming_header_column(cells):
    """
    柜类报价表头：返回「说明」列表头的 0-based 列索引。
    优先匹配同一行同时含「费用项目」与「说明」；否则取前若干行内首个单元格为「说明」的列。
    """
    if not cells:
        return None
    scan_rows = min(25, len(cells))
    for ri in range(scan_rows):
        row = cells[ri] or []
        labels = [str(x or "").strip() for x in row]
        if "费用项目" in labels and "说明" in labels:
            return labels.index("说明")
    for ri in range(scan_rows):
        row = cells[ri] or []
        for ci, val in enumerate(row):
            if str(val or "").strip() == "说明":
                return ci
    return None


def _dg_column_all_empty(cells, col_idx):
    if col_idx < 0:
        return True
    for row in cells or []:
        line = row or []
        if col_idx < len(line) and str(line[col_idx] or "").strip():
            return False
    return True


def trim_dg_grid_trailing_empty_cols_after_shuoming(cells, merges):
    """
    在「说明」列右侧去掉整列均为空的列（常见于 Excel 拖到 K 列但无内容）。
    找不到「说明」列时不修改。
    """
    if not cells:
        return cells, merges or []
    merges = merges or []
    desc_c = find_shuoming_header_column(cells)
    if desc_c is None:
        return cells, merges
    max_w = max((len(r) for r in cells if r), default=0)
    right = max_w
    while right > desc_c + 1 and _dg_column_all_empty(cells, right - 1):
        right -= 1
    if right >= max_w:
        return cells, merges
    out_cells = []
    for row in cells:
        r = list(row or [])[:right]
        while len(r) < right:
            r.append("")
        out_cells.append(r)
    out_merges = []
    for mg in merges:
        try:
            r0 = int(mg["r"])
            c0 = int(mg["c"])
            rs0 = int(mg["rs"])
            cs0 = int(mg["cs"])
        except (KeyError, TypeError, ValueError):
            continue
        if c0 >= right:
            continue
        end_c = c0 + cs0 - 1
        if end_c >= right:
            cs0 = right - c0
            if cs0 < 1:
                continue
        out_merges.append({"r": r0, "c": c0, "rs": rs0, "cs": cs0})
    return out_cells, out_merges


def _dg_remove_column_at(cells, merges, col_idx):
    """删除第 col_idx 列（0-based），同步调整 merges。"""
    new_cells = []
    for row in cells or []:
        r = list(row or [])
        if col_idx < len(r):
            r.pop(col_idx)
        new_cells.append(r)
    new_merges = []
    for mg in merges or []:
        try:
            r0 = int(mg["r"])
            c0 = int(mg["c"])
            rs0 = int(mg["rs"])
            cs0 = int(mg["cs"])
        except (KeyError, TypeError, ValueError):
            continue
        end_c = c0 + cs0 - 1
        if col_idx < c0:
            new_merges.append({"r": r0, "c": c0 - 1, "rs": rs0, "cs": cs0})
        elif col_idx > end_c:
            new_merges.append({"r": r0, "c": c0, "rs": rs0, "cs": cs0})
        else:
            new_cs = cs0 - 1
            if new_cs < 1:
                continue
            new_merges.append({"r": r0, "c": c0, "rs": rs0, "cs": new_cs})
    return new_cells, new_merges


def trim_fully_empty_columns(cells, merges):
    """
    去掉「每一行在该列都无有效内容」的列（含 Excel 中间/左侧留白），不限于「说明」右侧。
    合并单元格随删列同步收缩或左移。
    """
    if not cells:
        return cells or [], merges or []
    merges = list(merges or [])
    cells = [list(r) for r in cells]
    while True:
        max_w = max((len(r) for r in cells), default=0)
        if max_w <= 1:
            break
        for i in range(len(cells)):
            row = cells[i]
            if len(row) < max_w:
                row.extend([""] * (max_w - len(row)))
        victim = None
        for c in range(max_w - 1, -1, -1):
            if _dg_column_all_empty(cells, c):
                victim = c
                break
        if victim is None:
            break
        cells, merges = _dg_remove_column_at(cells, merges, victim)
    return cells, merges


def normalize_excel_grid_content_storage(content):
    """dg_grid 非 v2 保存入库前：裁全空列、横幅占位行、PS/免责行并入备注、整行首列合并推断，并刷新 rows/cols。"""
    if not isinstance(content, dict):
        return content
    if int(content.get("schema") or 0) == 2:
        return content
    cells = content.get("cells")
    if not isinstance(cells, list) or not cells:
        return content
    merges = content["merges"] if isinstance(content.get("merges"), list) else []
    cells = [list(r) for r in cells]
    cells, merges = trim_fully_empty_columns(cells, merges)
    cells, merges = strip_excel_quote_banner_gap_rows(cells, merges)
    remark = str(content.get('remark') or '')
    cells, merges, remark = strip_excel_quote_ps_disclaimer_rows(cells, merges, remark)
    content['remark'] = remark
    merges = _dg_merge_full_row_first_col_only(cells, merges)
    variant = str((content.get("variant") or "")).strip().lower()
    if variant == "excel_grid":
        hr_in = content.get("header_orange_row")
        cells, merges, hr_out = strip_interior_fully_empty_fee_rows(cells, merges, hr_in)
        content["header_orange_row"] = hr_out
    content["cells"] = cells
    content["merges"] = merges
    content["rows"] = len(cells)
    content["cols"] = len(cells[0]) if cells else 0
    return content


def canonical_excel_grid_module_content(content):
    """
    非 v2 栅格：与入库规范化一致的副本（裁列、横幅、PS 说明并入备注），不写库。
    schema v2 原样返回。
    """
    if not isinstance(content, dict):
        return {}
    if int(content.get('schema') or 0) == 2:
        return content
    out = copy.deepcopy(content)
    normalize_excel_grid_content_storage(out)
    return out


def _clip_excel_merge_to_output(mr, row_sources, c_lo, c_hi):
    """将 xlsx 合并区映射到去列后的网格（0-based，含 rs/cs）。"""
    mmin_r, mmin_c, mmax_r, mmax_c = mr.min_row, mr.min_col, mr.max_row, mr.max_col
    rows_in = [i for i, er in enumerate(row_sources) if mmin_r <= er <= mmax_r]
    if not rows_in:
        return None
    cc1 = max(mmin_c, c_lo)
    cc2 = min(mmax_c, c_hi)
    if cc1 > cc2:
        return None
    o_r1, o_r2 = min(rows_in), max(rows_in)
    # 输出网格列与 Excel 列对齐：保留区间从 c_lo 开始时，输出 0 列 = Excel 第 c_lo 列。
    o_c1, o_c2 = cc1 - c_lo, cc2 - c_lo
    rs, cs = o_r2 - o_r1 + 1, o_c2 - o_c1 + 1
    return {'r': o_r1, 'c': o_c1, 'rs': rs, 'cs': cs}


def _collect_output_merges(ws, row_sources, c_lo=_COL_KEEP_LO, c_hi=_COL_KEEP_HI):
    if not row_sources or not hasattr(ws, 'merged_cells') or not ws.merged_cells:
        return []
    out = []
    for mr in ws.merged_cells.ranges:
        item = _clip_excel_merge_to_output(mr, row_sources, c_lo, c_hi)
        if item and item['rs'] >= 1 and item['cs'] >= 1:
            out.append(item)
    return out


def _pad_cells(cells, n, m):
    if not cells:
        return [[''] * m for _ in range(n)]
    out = []
    for i in range(n):
        row = list(cells[i]) if i < len(cells) and cells[i] is not None else []
        if len(row) < m:
            row = row + [''] * (m - len(row))
        else:
            row = row[:m]
        out.append(row)
    return out


def prepare_dg_table_display(c):
    """
    与文章页 / dg_grid_render 过滤器前处理一致。
    入参 c 为模块 content 字典。返回 (cells, merges, header_orange_row)（尚未 pad/推断 merges 前的 merges 可仍为 None/[]）。
    schema≠2 且为 Excel 栅格时：裁掉「整列无内容」的空列再渲染（不必与 Excel 列宽一一对应）。
    """
    if c is None or not isinstance(c, dict):
        return ([[""]], None, None)
    cells = c.get("cells") or []
    merges_raw = c.get("merges")
    sch = int(c.get("schema") or 0)
    merges_work = list(merges_raw) if isinstance(merges_raw, list) else []
    if isinstance(merges_raw, list) and len(merges_raw) > 0:
        cells = [list(x) for x in cells]
    else:
        cells = normalize_dg_grid_cells(cells)
    if sch != 2 and cells:
        work_merges = list(merges_raw) if isinstance(merges_raw, list) else []
        work = {
            'schema': sch,
            'variant': c.get('variant'),
            'cells': [list(r) for r in cells],
            'merges': work_merges,
            'remark': str(c.get('remark') or ''),
            'header_orange_row': c.get('header_orange_row'),
        }
        normalize_excel_grid_content_storage(work)
        cells = work['cells']
        if isinstance(merges_raw, list):
            merges_work = work['merges']
        hr = work.get('header_orange_row')
    else:
        hr = c.get('header_orange_row')
    merges_out = merges_work if isinstance(merges_raw, list) else merges_raw
    if hr is not None and hr != '':
        try:
            hr = int(hr)
        except (TypeError, ValueError):
            hr = find_tidan_header_row(cells)
    else:
        hr = find_tidan_header_row(cells)
    return (cells, merges_out, hr)


def _prepare_dg_table_structure(cells, merges, header_orange_row):
    """
    Pad、按「仅首列非空行」补整行合并、建立 omit 矩阵。cells 为二维 list。
    返回 (cells, merges, n, m, omit, header_orange_row)（merges 为生效后的 list）。
    """
    cells = [list(r) for r in (cells or [])] if cells else [[]]
    n = len(cells)
    m = max((len(r) for r in cells if r), default=0) if n else 0
    if n < 1:
        m = max(m, 1)
        n = 1
    m = max(m, 1)
    cells = _pad_cells(cells, n, m)

    merges = list(merges or [])
    merges = _dg_merge_full_row_first_col_only(cells, merges)

    omit = [[False] * m for _ in range(n)]
    for mg in merges:
        r, c, rs, cs = int(mg["r"]), int(mg["c"]), int(mg["rs"]), int(mg["cs"])
        if r < 0 or c < 0 or r + rs > n or c + cs > m:
            continue
        for i in range(r, r + rs):
            for j in range(c, c + cs):
                if i < n and j < m:
                    omit[i][j] = True
        omit[r][c] = "anchor"

    if header_orange_row is None:
        header_orange_row = find_tidan_header_row(cells)
    return (cells, merges, n, m, omit, header_orange_row)


def split_dg_editable_label_value(text):
    """
    若文本适合拆成「标签」「值」两段（含中英文冒号），返回 (label, gap, value)；否则返回 None。
    gap 仅保留「至多一个空格」或为空（与 DG v2 键值间距一致）。
    """
    if text is None:
        return None
    s = normalize_dg_grid_cell_display_spaces(text)
    if not s:
        return None
    m = _DG_LV_SPLIT_RE.match(s)
    if not m:
        return None
    label, gap, value = m.group(1), m.group(2), m.group(3)
    if len(label.strip()) < 2:
        return None
    if "://" in label:
        return None
    # gap 只要存在空白分隔（即使仅为空格），保存时用单个半角空格；无分隔则不插入空格
    gap_norm = " " if gap else ""
    value_norm = re.sub(r"[\t\u3000 ]{2,}", " ", value.strip())
    return (label, gap_norm, value_norm)


def _dg_rest_of_row_empty(cells, r, skip_col, m):
    """第 r 行除 skip_col 外是否全为空（用于窄表仅 A 列填「标签：值」的行）。"""
    row = cells[r] if r < len(cells) else []
    for j in range(m):
        if j == skip_col:
            continue
        v = row[j] if j < len(row) else ""
        if str(v or "").strip():
            return False
    return True


def _dg_grid_editable_row_hide_delete_btn(cells, merges, r, m, omit):
    """
    Excel 栅格抬头区（截图主标题下「报价日期：」「装柜地址：」等）：不显示行删除按钮。
    条件：整行仅一格且 colspan 铺满，或仅首列有内容且其余列皆空（与 split 键值行一致）。
    """
    visible = []
    for c in range(m):
        op = omit[r][c]
        if op is True:
            continue
        mg = (
            next((g for g in merges if int(g["r"]) == r and int(g["c"]) == c), None)
            if op == "anchor"
            else None
        )
        cs2 = int(mg["cs"]) if mg else 1
        visible.append((c, op, cs2))
    if len(visible) == 0:
        return True
    if len(visible) == 1:
        c0, op0, cs2 = visible[0]
        if op0 == "anchor" and cs2 >= m:
            return True
        if c0 == 0 and cs2 == 1 and _dg_rest_of_row_empty(cells, r, 0, m):
            return True
    return False


def infer_merge_rows_only_first_col_nonempty(cells):
    """
    任意行：仅第 0 列有有效文本、第 1..m-1 列全空 → 推断整行横向合并 (0..m-1)。
    用于 Excel 导入后「键只在 A 列、右侧空列未合并」的抬头行，编辑端与前台合成一格。
    """
    if not cells:
        return []
    m = max((len(r) for r in cells if r), default=0)
    if m <= 1:
        return []
    out = []
    for r in range(len(cells)):
        row = cells[r]
        rline = list(row) + [''] * (m - len(row))
        if not str(rline[0] or '').strip():
            continue
        if not all(not str(rline[i] or '').strip() for i in range(1, m)):
            continue
        out.append({'r': r, 'c': 0, 'rs': 1, 'cs': m})
    return out


def _dg_merge_full_row_first_col_only(cells, merges):
    """
    对 infer_merge_rows_only_first_col_nonempty 命中的行：移除该行原有横向合并（rs==1），
    再写入整行 colspan=m，避免 Excel 未合并空列时出现多格与行删除按钮。
    """
    inferred = infer_merge_rows_only_first_col_nonempty(cells)
    if not inferred:
        return list(merges or [])
    merges = list(merges or [])
    for inf in inferred:
        r = int(inf['r'])
        merges = [
            mg for mg in merges
            if not (int(mg.get('r', 0)) == r and int(mg.get('rs', 1)) == 1)
        ]
        merges.append(inf)
    return merges


def infer_full_width_row_merges(cells):
    """兼容旧名：等价于「仅首列非空的行」整行合并推断（含原前两行标题逻辑）。"""
    return infer_merge_rows_only_first_col_nonempty(cells)


def render_dg_table_html(cells, merges=None, header_orange_row=None):
    """
    渲染与 xlsx 布局一致的 <table>（含 rowspan/colspan）；
    header_orange_row 为 0-based 行号时该行使用橙色表头底（与提单号行一致）。
    """
    raw = [list(r) for r in (cells or [])] if cells else [[]]
    cells, merges, n, m, omit, header_orange_row = _prepare_dg_table_structure(
        raw, merges, header_orange_row
    )

    parts = ['<table class="dg-grid-table dg-quote-v2-fee" role="table"><tbody>']
    for r in range(n):
        tr_cls = ' dg-grid-tr-tidan-header' if header_orange_row is not None and r == header_orange_row else ''
        parts.append(f'<tr class="dg-grid-tr{tr_cls}">')
        for c in range(m):
            op = omit[r][c]
            if op is True:
                continue
            val = cells[r][c] if c < len(cells[r]) else ''
            text = html.escape(normalize_dg_grid_cell_display_spaces(val), quote=True)
            if op == 'anchor':
                mg = next((g for g in merges if g['r'] == r and g['c'] == c), None)
                rs, cs2 = 1, 1
                if mg:
                    rs, cs2 = int(mg['rs']), int(mg['cs'])
                # 整行合并（如「中海豹国际」B2:J2）需取消 max-width，否则看起来像未合并
                full = ' dg-grid-cell--fullrow' if cs2 >= m else ''
                attr = f' class="dg-grid-cell{full}" rowspan="{rs}" colspan="{cs2}"'
            else:
                attr = ' class="dg-grid-cell"'
            parts.append(f'<td{attr}>{text}</td>')
        parts.append('</tr>')
    parts.append('</tbody></table>')
    return ''.join(parts)


def render_dg_table_editable_html(
    cells,
    merges=None,
    header_orange_row=None,
    *,
    data_r_base=0,
    table_extra_class="",
    include_row_action_column=True,
    readonly_row_set=None,
    fee_group_anchor_col=None,
    single_row_groups=False,
    group_header_rows=None,
    group_add_row_label="+ 费用",
    group_delete_label="删类别",
):
    """
    与 render_dg_table_html 同布局；可编辑单元格一般为单行 <input type="text">。
    若为整行合并且正文含「标签：值」，或首列为「标签：值」且该行其余列皆空，则拆成标签 / 值两个 input（保存时再拼回原字符串）。
    data_r_base：分段编辑器里费用子表的全局行号偏移。
    include_row_action_column：为 False 时不输出右侧删行列（抬头段不落占位 td，避免多出空白竖列）。
    readonly_row_set：行号集合（相对本表）；这些行 input 加 readonly + 灰底（用于表头行）。
    fee_group_anchor_col：在该列上的 rowspan(>1) 合并块视为「费用类别」，块的最末行在 row-action 列追加「+添加行」按钮。
    single_row_groups：True 时，anchor 列上非空的单行（无 rs>1 合并）也视为一个「主单」，同样在末行加「+费用」按钮；
        典型场景：Freightconn 类的 CODE + 主单 首列表，允许在单行主单下继续增加费用行。
    group_header_rows：行号集合；这些行是表头 / 横幅（例如 CODE 那行），即使 anchor 列非空也不视为主单。
    """
    raw = [list(r) for r in (cells or [])] if cells else [[]]
    cells, merges, n, m, omit, header_orange_row = _prepare_dg_table_structure(
        raw, merges, header_orange_row
    )
    ro_set = set(int(x) for x in (readonly_row_set or []))
    header_rows_set = set(int(x) for x in (group_header_rows or []))
    # 预计算「费用类别」rowspan 块的末行索引（按 fee_group_anchor_col 列上 rs>1 的合并）
    group_last_rows = set()
    group_meta = {}  # last_row -> (anchor_r, rs)
    if fee_group_anchor_col is not None:
        anchor_c = int(fee_group_anchor_col)
        for mg in merges:
            if int(mg.get('c', -1)) == anchor_c and int(mg.get('rs', 1)) > 1:
                r0 = int(mg['r'])
                rs0 = int(mg['rs'])
                last = r0 + rs0 - 1
                group_last_rows.add(last)
                group_meta[last] = (r0, rs0)
        if single_row_groups and 0 <= anchor_c < m:
            merged_anchor_rows = {int(mg['r']) for mg in merges
                                  if int(mg.get('c', -1)) == anchor_c and int(mg.get('rs', 1)) > 1}
            for r in range(n):
                if r in header_rows_set or r in ro_set:
                    continue
                if r in merged_anchor_rows:
                    continue
                if omit[r][anchor_c] is True:
                    continue
                val = cells[r][anchor_c] if anchor_c < len(cells[r]) else ''
                if not str(val or '').strip():
                    continue
                group_last_rows.add(r)
                group_meta[r] = (r, 1)

    tcls = "dg-grid-table dg-grid-table--editable dg-quote-v2-fee"
    if table_extra_class and str(table_extra_class).strip():
        tcls = f"{tcls} {str(table_extra_class).strip()}"
    parts = [f'<table class="{tcls}" role="table"><tbody>']
    for r in range(n):
        dr = r + int(data_r_base)
        tr_cls = ' dg-grid-tr-tidan-header' if header_orange_row is not None and r == header_orange_row else ''
        if r in ro_set:
            tr_cls += ' dg-grid-tr--readonly'
        if r in header_rows_set:
            tr_cls += ' dg-grid-tr--group-header'
        parts.append(f'<tr class="dg-grid-tr{tr_cls}">')
        ro_attr = ' readonly tabindex="-1"' if r in ro_set else ''
        ro_cls = ' dg-grid-cell-input--ro' if r in ro_set else ''
        for c in range(m):
            op = omit[r][c]
            if op is True:
                continue
            val = cells[r][c] if c < len(cells[r]) else ""
            val_clean = normalize_dg_grid_cell_display_spaces(val)
            val_esc = html.escape(val_clean, quote=True)
            lv = None
            mg = next((g for g in merges if g["r"] == r and g["c"] == c), None) if op == "anchor" else None
            rs, cs2 = 1, 1
            if mg:
                rs, cs2 = int(mg["rs"]), int(mg["cs"])
            del_grp = ""
            if (
                fee_group_anchor_col is not None
                and op == "anchor"
                and mg
                and int(mg["rs"]) > 1
                and c == int(fee_group_anchor_col)
            ):
                r_mg = int(mg["r"])
                g_drx = r_mg + int(data_r_base)
                fac = int(fee_group_anchor_col)
                del_grp = (
                    f'<button type="button" class="dg-fee-del-group-btn dg-fee-del-group-btn--merged-cell" '
                    f'data-dg-act="remove-fee-group" data-dg-group-anchor-r="{g_drx}" '
                    f'data-dg-group-anchor-c="{fac}" title="删除整个费用类别（本组合并行）"'
                    f' aria-label="删除本主单">'
                    f'<span class="dg-fee-del-group-btn__icon" aria-hidden="true">🗑</span>'
                    f'<span class="dg-fee-del-group-btn__text">{html.escape(group_delete_label, quote=True)}</span>'
                    f'</button>'
                )
            if op == "anchor":
                full = " dg-grid-cell--fullrow" if cs2 >= m else ""
                attr = f' class="dg-grid-cell{full}" rowspan="{rs}" colspan="{cs2}"'
                if cs2 >= m:
                    lv = split_dg_editable_label_value(val_clean)
            else:
                attr = ' class="dg-grid-cell"'
                if c == 0 and _dg_rest_of_row_empty(cells, r, 0, m):
                    lv = split_dg_editable_label_value(val_clean)
            if lv and r not in ro_set:
                lab, gap, vpart = lv
                lab_esc = html.escape(lab, quote=True)
                v_esc = html.escape(vpart, quote=True)
                sep_esc = html.escape(gap, quote=True)
                split_html = (
                    f'<span class="dg-grid-cell-split" data-r="{dr}" data-c="{c}" '
                    f'data-lv-sep="{sep_esc}">'
                    f'<input type="text" class="dg-grid-cell-input dg-grid-cell-input--lv '
                    f'dg-grid-cell-input--label" data-r="{dr}" data-c="{c}" data-part="label" '
                    f'spellcheck="false" autocomplete="off" value="{lab_esc}" />'
                    f'<input type="text" class="dg-grid-cell-input dg-grid-cell-input--lv '
                    f'dg-grid-cell-input--value" data-r="{dr}" data-c="{c}" data-part="value" '
                    f'spellcheck="false" autocomplete="off" value="{v_esc}" />'
                    f'</span>'
                )
                if del_grp:
                    parts.append(
                        f'<td{attr}><div class="dg-fee-td-merged-category">{split_html}{del_grp}</div></td>'
                    )
                else:
                    parts.append(f'<td{attr}>{split_html}</td>')
            else:
                inp_html = (
                    f'<input type="text" class="dg-grid-cell-input{ro_cls}" data-r="{dr}" data-c="{c}" '
                    f'spellcheck="false" autocomplete="off"{ro_attr} value="{val_esc}" />'
                )
                if del_grp:
                    parts.append(
                        f'<td{attr}><div class="dg-fee-td-merged-category">{inp_html}{del_grp}</div></td>'
                    )
                else:
                    parts.append(f'<td{attr}>{inp_html}</td>')
        if include_row_action_column:
            if r in ro_set or r in header_rows_set:
                parts.append(
                    '<td class="dg-grid-cell dg-grid-cell--row-action dg-grid-cell--row-action--static" '
                    'data-dg-side="actions" aria-hidden="true"></td>'
                )
            else:
                hide_del = _dg_grid_editable_row_hide_delete_btn(cells, merges, r, m, omit)
                extra_btn = ''
                if r in group_last_rows:
                    g_anchor_r, _g_rs = group_meta[r]
                    g_dr = g_anchor_r + int(data_r_base)
                    fac = int(fee_group_anchor_col)
                    extra_btn = (
                        f'<button type="button" class="dg-fee-add-row-btn" '
                        f'data-dg-act="append-row-in-group" data-dg-group-anchor-r="{g_dr}" '
                        f'data-dg-group-anchor-c="{fac}" title="在本主单下加费用行">{html.escape(group_add_row_label, quote=True)}</button>'
                    )
                if hide_del:
                    if extra_btn:
                        parts.append(
                            '<td class="dg-grid-cell dg-grid-cell--row-action" data-dg-side="actions">'
                            f'{extra_btn}</td>'
                        )
                    else:
                        parts.append(
                            '<td class="dg-grid-cell dg-grid-cell--row-action dg-grid-cell--row-action--static" '
                            'data-dg-side="actions" aria-hidden="true"></td>'
                        )
                else:
                    parts.append(
                        '<td class="dg-grid-cell dg-grid-cell--row-action" data-dg-side="actions">'
                        '<button type="button" class="dg-row-del-btn" data-dg-act="remove-row" title="删本行">'
                        f'删</button>{extra_btn}</td>'
                    )
        parts.append('</tr>')
    parts.append('</tbody></table>')
    return ''.join(parts)


def _render_dg_meta_editable_paired_html(mc, mm, data_r_base=0):
    """
    编辑器抬头段：KV 行（col0 含「标签：值」格式）两两配对在同一 <tr>；
    标题行（非 KV）独行全宽；奇数末尾 KV 行独行全宽。
    每个 input 保留原始 data-r/c，保存逻辑透明。
    """
    n = len(mc)
    if n == 0:
        return ''
    m = max((len(r) for r in mc if r), default=1)
    m = max(m, 1)
    padded = [list(r) + [''] * max(0, m - len(r)) for r in mc]

    def _is_kv(ri):
        col0 = normalize_dg_grid_cell_display_spaces(str(padded[ri][0] if padded[ri] else '')).strip()
        return bool(col0 and split_dg_editable_label_value(col0))

    def _cell_inner(ri, val_clean):
        dr = ri + data_r_base
        lv = split_dg_editable_label_value(val_clean)
        if lv:
            lab, gap, vpart = lv
            le = html.escape(lab, quote=True)
            ve = html.escape(vpart, quote=True)
            se = html.escape(gap, quote=True)
            return (
                f'<span class="dg-grid-cell-split" data-r="{dr}" data-c="0" data-lv-sep="{se}">'
                f'<input type="text" class="dg-grid-cell-input dg-grid-cell-input--lv dg-grid-cell-input--label" '
                f'data-r="{dr}" data-c="0" data-part="label" spellcheck="false" autocomplete="off" value="{le}" />'
                f'<input type="text" class="dg-grid-cell-input dg-grid-cell-input--lv dg-grid-cell-input--value" '
                f'data-r="{dr}" data-c="0" data-part="value" spellcheck="false" autocomplete="off" value="{ve}" />'
                f'</span>'
            )
        ve = html.escape(val_clean, quote=True)
        return (
            f'<input type="text" class="dg-grid-cell-input" data-r="{dr}" data-c="0" '
            f'spellcheck="false" autocomplete="off" value="{ve}" />'
        )

    # rs>1 的 rowspan 后续行才需要 omit（单行合并不影响）
    omit = [False] * n
    for mg in mm:
        r0, rs = int(mg.get('r', 0)), int(mg.get('rs', 1))
        if rs > 1:
            for ki in range(r0 + 1, min(r0 + rs, n)):
                omit[ki] = True

    parts = [
        '<table class="dg-grid-table dg-grid-table--editable dg-quote-v2-fee '
        'dg-grid-table--excel-meta" role="table">'
        '<colgroup><col style="width:50%"><col style="width:50%"></colgroup>'
        '<tbody>'
    ]
    i = 0
    while i < n:
        if omit[i]:
            i += 1
            continue
        val0 = normalize_dg_grid_cell_display_spaces(str(padded[i][0] or ''))
        if _is_kv(i):
            j = i + 1
            while j < n and omit[j]:
                j += 1
            if j < n and _is_kv(j):
                val1 = normalize_dg_grid_cell_display_spaces(str(padded[j][0] or ''))
                parts.append('<tr class="dg-grid-tr dg-meta-kv-paired">')
                parts.append(f'<td class="dg-grid-cell dg-meta-kv-half">{_cell_inner(i, val0)}</td>')
                parts.append(f'<td class="dg-grid-cell dg-meta-kv-half">{_cell_inner(j, val1)}</td>')
                parts.append('</tr>')
                omit[j] = True
                i += 1
                continue
        parts.append('<tr class="dg-grid-tr">')
        parts.append(
            f'<td class="dg-grid-cell dg-grid-cell--fullrow" colspan="2">'
            f'{_cell_inner(i, val0)}</td>'
        )
        parts.append('</tr>')
        i += 1
    parts.append('</tbody></table>')
    return ''.join(parts)


def render_dg_fee_div_editor_html(
    cells,
    merges=None,
    header_orange_row=None,
    *,
    data_r_base=0,
    readonly_row_set=None,
    fee_group_anchor_col=0,
):
    """
    Excel 费用子表编辑器：div + CSS Grid（非 <table>），左侧大类纵向合并与右侧细项同表格语义；
    依赖 cells + merges 存库；与 render_dg_table_editable_html 共用 omit/合并矩阵。
    """
    raw = [list(r) for r in (cells or [])] if cells else [[]]
    cells, merges, n, m, omit, header_orange_row = _prepare_dg_table_structure(
        raw, merges, header_orange_row
    )
    ro_set = set(int(x) for x in (readonly_row_set or []))
    anchor_c = int(fee_group_anchor_col)
    group_last_rows = set()
    group_meta = {}
    for mg in merges:
        if int(mg.get("c", -1)) == anchor_c and int(mg.get("rs", 1)) > 1:
            r0 = int(mg["r"])
            rs0 = int(mg["rs"])
            last = r0 + rs0 - 1
            group_last_rows.add(last)
            group_meta[last] = (r0, rs0)

    base = int(data_r_base)

    def emit_cell(rr, c, gc, gr):
        dr = rr + base
        val = cells[rr][c] if c < len(cells[rr]) else ""
        val_clean = normalize_dg_grid_cell_display_spaces(val)
        val_esc = html.escape(val_clean, quote=True)
        ro_attr = ' readonly tabindex="-1"' if rr in ro_set else ""
        ro_cls = " dg-grid-cell-input--ro" if rr in ro_set else ""
        op = omit[rr][c]
        mg = (
            next((g for g in merges if int(g["r"]) == rr and int(g["c"]) == c), None)
            if op == "anchor"
            else None
        )
        rs_m, cs_m = (int(mg["rs"]), int(mg["cs"])) if mg else (1, 1)
        lv = None
        if op == "anchor" and cs_m >= m:
            lv = split_dg_editable_label_value(val_clean)
        elif op != "anchor" and c == 0 and _dg_rest_of_row_empty(cells, rr, 0, m):
            lv = split_dg_editable_label_value(val_clean)
        tidan_cls = (
            " dg-fee-div-cell--tidan"
            if header_orange_row is not None and rr == header_orange_row
            else ""
        )
        if lv and rr not in ro_set:
            lab, gap, vpart = lv
            lab_esc = html.escape(lab, quote=True)
            v_esc = html.escape(vpart, quote=True)
            sep_esc = html.escape(gap, quote=True)
            inner = (
                f'<span class="dg-grid-cell-split" data-r="{dr}" data-c="{c}" data-lv-sep="{sep_esc}">'
                f'<input type="text" class="dg-grid-cell-input dg-grid-cell-input--lv '
                f'dg-grid-cell-input--label" data-r="{dr}" data-c="{c}" data-part="label" '
                f'spellcheck="false" autocomplete="off" value="{lab_esc}" />'
                f'<input type="text" class="dg-grid-cell-input dg-grid-cell-input--lv '
                f'dg-grid-cell-input--value" data-r="{dr}" data-c="{c}" data-part="value" '
                f'spellcheck="false" autocomplete="off" value="{v_esc}" />'
                "</span>"
            )
        else:
            inner = (
                f'<input type="text" class="dg-grid-cell-input{ro_cls}" data-r="{dr}" '
                f'data-c="{c}" spellcheck="false" autocomplete="off"{ro_attr} value="{val_esc}" />'
            )
        merged_cat_cls = ""
        del_cat_btn = ""
        if (
            op == "anchor"
            and mg
            and int(mg["rs"]) > 1
            and int(c) == anchor_c
        ):
            g_dr = rr + base
            del_cat_btn = (
                f'<button type="button" class="dg-fee-del-group-btn dg-fee-del-group-btn--merged-cell" '
                f'data-dg-act="remove-fee-group" data-dg-group-anchor-r="{g_dr}" '
                f'data-dg-group-anchor-c="{anchor_c}" title="删除整个费用类别（本组合并行）">删类别</button>'
            )
            merged_cat_cls = " dg-fee-div-cell--merged-category"
        body = (
            f'<div class="dg-fee-div-merged-category">{inner}{del_cat_btn}</div>'
            if del_cat_btn
            else inner
        )
        return (
            f'<div class="dg-fee-div-cell{tidan_cls}{merged_cat_cls}" '
            f'style="grid-column:{gc} / span {cs_m};grid-row:{gr} / span {rs_m}">{body}</div>'
        )

    def emit_action(rr, gr):
        dr = rr + base
        tidan_act = (
            " dg-fee-div-cell--tidan"
            if header_orange_row is not None and rr == header_orange_row
            else ""
        )
        hide_del = _dg_grid_editable_row_hide_delete_btn(cells, merges, rr, m, omit)
        extra_btn = ""
        if rr in group_last_rows:
            ar0, _rs = group_meta[rr]
            g_dr = ar0 + base
            extra_btn = (
                f'<button type="button" class="dg-fee-add-row-btn" '
                f'data-dg-act="append-row-in-group" data-dg-group-anchor-r="{g_dr}" '
                f'data-dg-group-anchor-c="{anchor_c}" title="本组添加行">+行</button>'
            )
        if hide_del:
            inner = extra_btn if extra_btn else '<span class="dg-fee-div-action-spacer" aria-hidden="true"></span>'
        else:
            inner = (
                '<button type="button" class="dg-row-del-btn" data-dg-act="remove-row" '
                'title="删本行">删</button>' + extra_btn
            )
        return (
            f'<div class="dg-fee-div-cell dg-fee-div-cell--action{tidan_act}" '
            f'data-dg-abs-r="{dr}" style="grid-column:{m + 1};grid-row:{gr} / span 1">{inner}</div>'
        )

    parts = [
        f'<div class="dg-fee-div-editor dg-quote-v2-fee" data-dg-fee-div="1" '
        f'data-dg-fee-cols="{m}" data-dg-fee-anchor-col="{anchor_c}">'
    ]
    parts.append(
        f'<div class="dg-fee-div-head dg-fee-div-row-grid dg-fee-div-row--readonly" '
        f'style="display:grid;grid-template-columns:repeat({m},minmax(0,1fr)) 52px;">'
    )
    for c in range(m):
        op = omit[0][c]
        if op is True:
            continue
        parts.append(emit_cell(0, c, c + 1, 1))
    parts.append(emit_action(0, 1))
    parts.append("</div>")

    parts.append('<div class="dg-fee-div-body">')
    r = 1
    while r < n:
        mg0 = next((g for g in merges if int(g["r"]) == r and int(g["c"]) == anchor_c), None)
        if mg0 and int(mg0["rs"]) > 1:
            rs0 = int(mg0["rs"])
            r0 = r
            g_dr_anchor = r0 + base
            parts.append(
                f'<div class="dg-fee-div-group" data-dg-group-anchor-r="{g_dr_anchor}" '
                f'data-dg-group-anchor-c="{anchor_c}">'
            )
            parts.append(
                f'<div class="dg-fee-div-group-grid" style="display:grid;grid-template-columns:'
                f'repeat({m},minmax(0,1fr)) 52px;grid-template-rows:'
                f'repeat({rs0},minmax(36px,auto));">'
            )
            for rr in range(r0, r0 + rs0):
                gr = rr - r0 + 1
                for c in range(m):
                    op = omit[rr][c]
                    if op is True:
                        continue
                    parts.append(emit_cell(rr, c, c + 1, gr))
                parts.append(emit_action(rr, gr))
            parts.append("</div></div>")
            r += rs0
        else:
            parts.append('<div class="dg-fee-div-standalone">')
            parts.append(
                f'<div class="dg-fee-div-row-grid" style="display:grid;grid-template-columns:'
                f'repeat({m},minmax(0,1fr)) 52px;">'
            )
            rr = r
            for c in range(m):
                op = omit[rr][c]
                if op is True:
                    continue
                parts.append(emit_cell(rr, c, c + 1, 1))
            parts.append(emit_action(rr, 1))
            parts.append("</div></div>")
            r += 1
    parts.append("</div></div>")
    return "".join(parts)


def render_dg_excel_grid_editable_split_html(cells, merges, header_orange_row_full):
    """
    后台 Excel 栅格编辑器：抬头一块表（KV 行两两配对）+ 「费用项目」版块标题 + 费用明细表。
    """
    fee_hdr = find_excel_quote_fee_header_row(cells)
    if fee_hdr is None:
        return render_dg_table_editable_html(cells, merges, header_orange_row_full)

    meta_cells = cells[:fee_hdr]
    fee_cells = cells[fee_hdr:]
    meta_merges = [mg for mg in (merges or []) if int(mg.get("r", 0)) < fee_hdr]
    fee_merges = _fee_slice_merges(merges, fee_hdr)
    fee_hr = _fee_relative_header_orange_row(header_orange_row_full, fee_hdr, fee_cells)

    mc = [list(r) for r in meta_cells]
    mm = [dict(mg) for mg in meta_merges]
    mc, mm = trim_fully_empty_columns(mc, mm)

    parts = ['<div class="dg-excel-editor-split">']
    parts.append('<div class="dg-quote-v2-section dg-excel-editor-meta-block">')
    parts.append('<div class="dg-quote-v2-fee-wrap dg-excel-editor-meta-wrap">')
    parts.append(_render_dg_meta_editable_paired_html(mc, mm, data_r_base=0))
    parts.append('</div></div>')
    parts.append('<h3 class="dg-quote-v2-sec-title dg-excel-editor-fee-heading">费用项目</h3>')
    parts.append('<div class="dg-quote-v2-section dg-excel-editor-fee-block dg-excel-editor-fee-sheet">')
    parts.append('<div class="dg-quote-v2-fee-wrap dg-excel-editor-fee-wrap">')
    parts.append(
        render_dg_fee_div_editor_html(
            fee_cells,
            fee_merges,
            fee_hr,
            data_r_base=fee_hdr,
            readonly_row_set={0},
            fee_group_anchor_col=0,
        )
    )
    parts.append('</div></div></div>')
    return ''.join(parts)


# --- DG 模块 schema v2：对齐「4月25日DG周鹏.xlsx」结构（上：客户与单证，下：费用表）---
DG_SCHEMA_V2 = 2


def default_dg_v2_customer_rows():
    """与项目根目录「4月25日DG周鹏.xlsx」Sheet1 客户区一致（5 行 × 两列键值对）。"""
    return [
        {"label": "客户ID编码：", "value": "周鹏"},
        {"label": "报价日期：", "value": "2026-04-25"},
        {"label": "客户地址：", "value": ""},
        {
            "label": "公司地址：",
            "value": "深圳市宝安区福永街道翠岗西路怀德国际大厦1801",
        },
        {"label": "联系人/职位：", "value": ""},
        {"label": "联系人：", "value": ""},
        {"label": "电话/分机：", "value": ""},
        {"label": "电话：", "value": ""},
        {"label": "Email：", "value": ""},
        {"label": "Email：", "value": ""},
    ]


def default_dg_v2_logistics():
    """起运/目的/船名；提单号与柜号在费用表各列，不在此表。与 4月25日DG周鹏.xlsx 起运/单证区一致。"""
    return {
        "origin_port": "深圳",
        "transit_port": "",
        "dest_port": "SYD海外仓",
        "dest_warehouse": "",
        "vessel": "",
    }


def _dg_v2_fee_subtotal(name, amount, note=""):
    """与 xlsx 一致：「费用项目」+「币别」合并为同一格（row_kind=subtotal）。"""
    amt = ""
    if amount is not None and amount != "":
        amt = str(amount)
    return {
        "row_kind": "subtotal",
        "bl_no": "",
        "container_no": "",
        "name": name,
        "currency": "",
        "amount": amt,
        "note": note or "",
    }


def default_dg_v2_fee_items():
    """与 4月25日DG周鹏.xlsx 费用行一致（含 RMB 小计、汇率说明行、总合计，合并格同表）。"""
    z = {"row_kind": "item", "bl_no": "", "container_no": ""}
    return [
        {
            "name": "Container loading fee/装箱费",
            "currency": "RMB",
            "amount": "800",
            "note": "",
            **z,
        },
        {
            "name": "customs clearance fee/报关费",
            "currency": "RMB",
            "amount": "500",
            "note": "",
            **z,
        },
        {
            "name": "Maritime declaration fee/海事申报费",
            "currency": "RMB",
            "amount": "不确定项",
            "note": "400/个un",
            **z,
        },
        {
            "name": "supervision fee for loading/危险品监装费",
            "currency": "RMB",
            "amount": "不确定项",
            "note": "低消1500，超出550/小时",
            **z,
        },
        {
            "name": "container stuffing charge/提箱费",
            "currency": "RMB",
            "amount": "2000",
            "note": "",
            **z,
        },
        _dg_v2_fee_subtotal("RMB合计", 3300),
        {
            "name": "B/L surrender fee/电放费",
            "currency": "RMB",
            "amount": "450",
            "note": "",
            **z,
        },
        {
            "name": "Sealing fee/封条费",
            "currency": "RMB",
            "amount": "50",
            "note": "",
            **z,
        },
        {
            "name": "THC/码头操作费",
            "currency": "RMB",
            "amount": "900",
            "note": "",
            **z,
        },
        {
            "name": "DOC/文件",
            "currency": "RMB",
            "amount": "450",
            "note": "",
            **z,
        },
        {
            "name": "ocean freight/海运费",
            "currency": "USD",
            "amount": "1750",
            "note": "20GP",
            **z,
        },
        _dg_v2_fee_subtotal("RMB合计", 14450),
        {
            "name": "码头杂费",
            "currency": "AUD",
            "amount": "850",
            "note": "",
            **z,
        },
        {
            "name": "DOC/文件费",
            "currency": "AUD",
            "amount": "300",
            "note": "",
            **z,
        },
        {
            "name": "Custom Clearance清关费+税金",
            "currency": "AUD",
            "amount": "180",
            "note": "税金实报实销",
            **z,
        },
        {
            "name": "Devanning/拆柜",
            "currency": "AUD",
            "amount": "",
            "note": "",
            **z,
        },
        {
            "name": "Delivery fee/派送费",
            "currency": "AUD",
            "amount": "1100",
            "note": "",
            **z,
        },
        _dg_v2_fee_subtotal("RMB合计", 11907),
        _dg_v2_fee_subtotal("美金/澳币兑换人民币汇率", "", "参考即时汇率"),
        _dg_v2_fee_subtotal("总合计", 29657, "不包含不确定项"),
    ]


def is_dg_content_v2(c):
    """新结构：schema=2。旧数据无此字段，仍为合并单元格矩阵。"""
    if not c or not isinstance(c, dict):
        return False
    try:
        return int(c.get("schema") or 0) == DG_SCHEMA_V2
    except (TypeError, ValueError):
        return False


def merge_dg_v2_content(d):
    """补全 v2 缺省键，供后台/渲染使用。"""
    d = dict(d) if d else {}
    if not d.get("customer_rows") or not isinstance(d.get("customer_rows"), list):
        d["customer_rows"] = default_dg_v2_customer_rows()
    else:
        # 与默认条数不一致时只保留已有，不强制补齐
        pass
    if not d.get("logistics") or not isinstance(d.get("logistics"), dict):
        d["logistics"] = default_dg_v2_logistics()
    else:
        base = default_dg_v2_logistics()
        base.update(d["logistics"])
        d["logistics"] = base
    # 旧数据：提单/柜在 logistics 时迁入费用行
    mbl = (d.get("logistics") or {}).get("bl_no") or ""
    mct = (d.get("logistics") or {}).get("container_no") or ""
    if "logistics" in d and isinstance(d["logistics"], dict):
        d["logistics"].pop("bl_no", None)
        d["logistics"].pop("container_no", None)
    if not d.get("fee_items") or not isinstance(d.get("fee_items"), list):
        d["fee_items"] = default_dg_v2_fee_items()
    elif len(d["fee_items"]) == 0:
        d["fee_items"] = default_dg_v2_fee_items()
    any_bl = any(
        (isinstance(x, dict) and (x.get("bl_no") or "").strip()) for x in (d.get("fee_items") or [])
    )
    any_cn = any(
        (isinstance(x, dict) and (x.get("container_no") or "").strip()) for x in (d.get("fee_items") or [])
    )
    for it in d.get("fee_items") or []:
        if not isinstance(it, dict):
            continue
        it.setdefault("bl_no", "")
        it.setdefault("container_no", "")
        it.setdefault("amount_alt", "")
        rk = it.get("row_kind") or "item"
        it["row_kind"] = rk if rk in ("item", "subtotal") else "item"
        if it["row_kind"] == "subtotal":
            it["currency"] = ""
        if it["row_kind"] != "subtotal":
            if mbl and not any_bl and not (it.get("bl_no") or "").strip():
                it["bl_no"] = mbl
            if mct and not any_cn and not (it.get("container_no") or "").strip():
                it["container_no"] = mct
    dual_fee = bool(d.get("fee_dual_amount"))
    d["fee_dual_amount"] = dual_fee
    if dual_fee:
        d.setdefault("fee_amount_label_1", "20GP/金额")
        d.setdefault("fee_amount_label_2", "40HQ/金额")
    d.setdefault("header_main", "中海豹国际")
    d.setdefault("header_sub", "报价单")
    d.pop("footer_note", None)
    d.setdefault("title", "DG报价表")
    d.setdefault("template_file", "")
    d.setdefault("sheet_name", "")
    d.setdefault("remark", "")
    d["schema"] = DG_SCHEMA_V2
    return normalize_dg_v2_ps_rows_to_remark(d)


def render_dg_quote_v2_html(c):
    """文章页 v2 模块 HTML（抬头+客户+单证与 Excel 同一张四列表，不含模块标题与 remark 外层）。"""
    from markupsafe import Markup, escape

    c = merge_dg_v2_content(c if isinstance(c, dict) else {})
    parts = ['<div class="dg-quote-v2">']
    # 与 xlsx 一致：大标题、副标题、客户两列一组、起运/单证两列一组，合并在同一 <table> 内
    parts.append(
        '<table class="dg-quote-v2-meta" role="presentation" aria-label="报价抬头与客户及单证">'
    )
    parts.append(
        f'<tr class="dg-c-main"><th colspan="4" scope="col">{escape(c.get("header_main") or "")}</th></tr>'
    )
    parts.append(
        f'<tr class="dg-c-sub"><th colspan="4" scope="col">{escape(c.get("header_sub") or "")}</th></tr>'
    )
    cust = c.get("customer_rows") or []
    for i in range(0, len(cust), 2):
        a, b = cust[i], (cust[i + 1] if i + 1 < len(cust) else None)
        parts.append(
            f'<tr class="dg-c-kv">'
            f'<th scope="row" class="dg-c-lab">{escape(str(a.get("label") or ""))}</th>'
        )
        if b is not None:
            parts.append(
                f'<td class="dg-c-val">{escape(str(a.get("value") or ""))}</td>'
                f'<th scope="row" class="dg-c-lab">{escape(str(b.get("label") or ""))}</th>'
                f'<td class="dg-c-val">{escape(str(b.get("value") or ""))}</td></tr>'
            )
        else:
            parts.append(
                f'<td class="dg-c-val dg-c-val--wide" colspan="3">'
                f'{escape(str(a.get("value") or ""))}</td></tr>'
            )
    log = c.get("logistics") or {}
    log_quads = [
        (
            ("起运港", "origin_port"),
            ("中转港", "transit_port"),
        ),
        (
            ("目的港", "dest_port"),
            ("海外仓", "dest_warehouse"),
        ),
    ]
    for quad in log_quads:
        left, right = quad[0], quad[1]
        parts.append(
            f'<tr class="dg-c-log">'
            f'<th scope="row" class="dg-c-lab">{escape(left[0])}</th>'
            f'<td class="dg-c-val">{escape(str(log.get(left[1]) or ""))}</td>'
            f'<th scope="row" class="dg-c-lab">{escape(right[0])}</th>'
            f'<td class="dg-c-val">{escape(str(log.get(right[1]) or ""))}</td></tr>'
        )
    # 提单号、柜号在下方费用表各列
    parts.append(
        f'<tr class="dg-c-log">'
        f'<th scope="row" class="dg-c-lab">船名航次</th>'
        f'<td class="dg-c-val dg-c-val--wide" colspan="3">'
        f'{escape(str(log.get("vessel") or ""))}</td></tr>'
    )
    parts.append("</table>")
    dual = bool(c.get("fee_dual_amount"))
    fee_cls = "dg-quote-v2-fee" + (" dg-quote-v2-fee--dual" if dual else "")
    lab1 = escape(str(c.get("fee_amount_label_1") or "20GP/金额"))
    lab2 = escape(str(c.get("fee_amount_label_2") or "40HQ/金额"))
    parts.append(
        '<div class="dg-quote-v2-section dg-quote-v2-fee-block">'
        '<h3 class="dg-quote-v2-sec-title">费用项目</h3>'
        f'<div class="dg-quote-v2-fee-wrap"><table class="{fee_cls}">'
    )
    if dual:
        parts.append(
            f"<thead><tr><th>提单号</th><th>柜号</th><th>费用项目</th><th>币别</th>"
            f"<th>{lab1}</th><th>{lab2}</th><th>说明</th></tr></thead><tbody>"
        )
    else:
        parts.append(
            "<thead><tr><th>提单号</th><th>柜号</th><th>费用项目</th><th>币别</th>"
            "<th>金额</th><th>说明</th></tr></thead><tbody>"
        )
    for it in c.get("fee_items") or []:
        is_sub = (it.get("row_kind") or "item") == "subtotal"
        cl = "dg-quote-v2-fee-tr" + (
            " dg-quote-v2-fee-tr--subtotal" if is_sub else " dg-quote-v2-fee-tr--item"
        )
        parts.append(f'<tr class="{cl}">')
        if is_sub:
            parts.append(f"<td>{escape(str(it.get('bl_no') or ''))}</td>")
            parts.append(f"<td>{escape(str(it.get('container_no') or ''))}</td>")
            parts.append(
                f'<td colspan="2" class="dg-c-fee-merged-lab">'
                f"{escape(str(it.get('name') or ''))}</td>"
            )
            parts.append(f"<td>{escape(str(it.get('amount') or ''))}</td>")
            if dual:
                parts.append(f"<td>{escape(str(it.get('amount_alt') or ''))}</td>")
            parts.append(f"<td>{escape(str(it.get('note') or ''))}</td>")
        else:
            keys = (
                ("bl_no", "container_no", "name", "currency", "amount", "amount_alt", "note")
                if dual
                else ("bl_no", "container_no", "name", "currency", "amount", "note")
            )
            for k in keys:
                parts.append(f"<td>{escape(str(it.get(k) or ''))}</td>")
        parts.append("</tr>")
    parts.append("</tbody></table></div></div>")
    parts.append("</div>")
    return Markup("".join(parts))


def empty_dg_grid_content():
    """
    新建「DG报价表」：schema v2，结构对齐 4月25日DG周鹏.xlsx（客户信息 + 费用表，可增删行）。
    """
    return {
        "schema": DG_SCHEMA_V2,
        "variant": "dg_quote_v2",
        "title": "DG报价表",
        "header_main": "中海豹国际",
        "header_sub": "报价单",
        "customer_rows": default_dg_v2_customer_rows(),
        "logistics": default_dg_v2_logistics(),
        "fee_items": default_dg_v2_fee_items(),
        "template_file": "",
        "sheet_name": "",
        "remark": (
            "PS:如果遇到海关查验，查验费用实报实销。\n\n"
            "备注：2620"
        ),
    }


def empty_dg_excel_grid_content(title=None, template_id=None):
    """
    新建 Excel 严格栅格空白模板（如 9类电池柜报价表 / 普柜报价表）；导入 xlsx 见 cabinet_quote_xlsx_grid_content。
    """
    tid = (template_id or "").strip()
    t = (title or "").strip()
    if not t:
        if tid == "普柜":
            t = "普柜报价表"
        elif tid == "9类电池柜":
            t = "9类电池柜报价表"
        else:
            t = "9类电池柜报价表"
    return {
        "variant": "excel_grid",
        "template_id": tid,
        "title": t,
        "template_file": "",
        "sheet_name": "",
        "rows": 1,
        "cols": 1,
        "cells": [[""]],
        "merges": [],
        "header_orange_row": None,
        "remark": "",
    }


def cabinet_quote_xlsx_grid_content(path):
    """
    柜类报价：从 xlsx 第一工作表导入为 DG **网格**（cells + merges）。
    导入后去掉「全列无内容」的空列（不限于说明列右侧），合并区同步收缩；展示以版面为准，不必保留 Excel 冗余空列。
    """
    data = build_dg_grid_content_from_xlsx(path, title="")
    data.pop("schema", None)
    abs_path = path if os.path.isabs(path) else os.path.join(config._BASE_DIR, path)
    stem = os.path.splitext(os.path.basename(abs_path))[0]
    data["title"] = f"{stem}报价表"
    data["variant"] = "excel_grid"
    data["template_id"] = stem
    data.setdefault("remark", "")
    normalize_excel_grid_content_storage(data)
    return data


def build_dg_grid_content_from_xlsx(path, title=None):
    """
    从任意 xlsx 第一页构建 DG 网格模块 content（cells + merges）。
    宽表（列数 ≥ DG_EDGE_TRIM_MIN_COLS）仍按原逻辑裁左右列并只对 B–J 采集合并；
    窄表保留全部列，合并从 A 列映射（柜类报价见 cabinet_quote_xlsx_grid_content）。
    """
    if load_workbook is None:
        return _fallback_with_meta('未安装 openpyxl，请执行 pip install -r requirements.txt')

    abs_path = path if os.path.isabs(path) else os.path.join(config._BASE_DIR, path)
    if not os.path.isfile(abs_path):
        return _fallback_with_meta(f'未找到文件：{os.path.basename(abs_path)}')

    wb = load_workbook(abs_path, read_only=False, data_only=True)
    try:
        ws = wb.worksheets[0]
        sheet_name, _nr, _nc, cells, row_src = _read_sheet_grid_with_sources(ws)
        cells, row_src = trim_leading_empty_rows_paired(cells, row_src)
        max_w = max((len(r) for r in cells if r), default=0)
        if max_w >= DG_EDGE_TRIM_MIN_COLS:
            merges_src = _collect_output_merges(ws, row_src, _COL_KEEP_LO, _COL_KEEP_HI)
            cells = trim_dg_edge_columns(cells)
        else:
            mc = max(int(ws.max_column or 1), 1)
            merges_src = _collect_output_merges(ws, row_src, 1, mc)
    finally:
        wb.close()

    nrows = len(cells)
    ncols = len(cells[0]) if cells and cells[0] is not None else 0
    merges = [m for m in merges_src if m['c'] + m['cs'] <= ncols and m['r'] + m['rs'] <= nrows]
    header_row = find_tidan_header_row(cells)
    base_name = os.path.basename(abs_path)
    stem = os.path.splitext(base_name)[0]
    return {
        'title': stem if title is None else title,
        'template_file': base_name,
        'sheet_name': sheet_name,
        'rows': nrows,
        'cols': ncols,
        'cells': cells,
        'merges': merges,
        'header_orange_row': header_row,
        'remark': '',
    }


def build_default_dg_grid_content():
    """读取 config 默认 xlsx 第一页（与 DG_QUOTE_XLSX_PATH）。"""
    return build_dg_grid_content_from_xlsx(_abs_path(), title='4月25日 · DG 周鹏报价表')


def _fallback_with_meta(hint=''):
    return {
        'title': '报价表',
        'template_file': '',
        'sheet_name': '',
        'rows': 1,
        'cols': 1,
        'cells': [[hint or '无数据']],
        'merges': [],
        'header_orange_row': None,
        'remark': '',
    }


def fallback_empty_grid(hint=''):
    return _fallback_with_meta(hint)
